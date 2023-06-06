from pydcc_tables import DccTabel, DccTableColumn
import openpyxl as pyxl
import shutil

colAttrDefs = ("scopeType", "columnType", "measurandType", "unit", "humanHeading")
tblAttrDefs =("tableID", "itemID", "numRows", "numColumns")

def transpose_2d_list(matrix):
    return [list(row) for row in zip(*matrix)]

def _test_get_tables_from_sheet(sheetName="Table2"):
    """ Function that finds all the tables in a given sheet """

    wb = pyxl.load_workbook("DCC-Table_example3.xlsx", data_only=True)

    ws = wb[sheetName]


    columns = []

    tableID = ws["B2"].value
    itemID = ws["B3"].value
    numRows = ws["B4"].value
    numColumns = ws["B5"].value

    nRows = int(numRows)+5
    nCols = int(numColumns)

    cell = ws["B6"]

    content = [[cell.offset(r,c).value for r in range(nRows)] for c in range(nCols)]
    # content = transpose_2d_list(content)

    for c in content:
        col = DccTableColumn(   scopeType=c[0],
                                columnType=c[1],
                                measurandType=c[2],
                                unit=c[3],
                                humanHeading = c[4],
                                columnData= list(map(str, c[5:])))
        columns.append(col)

    tbl = DccTabel(tableID, itemID, numRows, numColumns, columns)
    wb.close()
    return tbl

def write_DCC_table_to_excel_sheet(dccTbl: DccTabel, workbookName = ""):
    tbl = dccTbl
    shutil.copy("DCC-Table_empty_template.xlsx", workbookName)
    wb = pyxl.load_workbook(workbookName)
    ws = wb["TableTemplate"]
    newws = wb.copy_worksheet(ws)
    newws.title = dccTbl.tableID
    ws = wb[dccTbl.tableID]
    wb.active = wb[dccTbl.tableID]
    print(wb.sheetnames)

    ws["B2"] = tbl.tableID
    ws["B3"] = tbl.itemID
    ws["B4"] = tbl.numRows
    ws["B5"] = tbl.numColumns

    cell = ws["B6"]
    columns = tbl.columns

    for c in range(tbl.numColumns):
        for r in range(len(colAttrDefs)):
            cell.offset(r, c).value = getattr(columns[c], colAttrDefs[r])
        for r in range(tbl.numRows):
            cell.offset(r+5, c).value = columns[c].columnData[r]

    wb.save(workbookName)



if __name__ == "__main__":
    tbl = _test_get_tables_from_sheet()
    columns = tbl.columns
    columns[5].print()
    for i in range(tbl.numColumns):
        print(columns[i].columnData)
    write_DCC_table_to_excel_sheet(tbl, "DCC-Table_example_output.xlsx")