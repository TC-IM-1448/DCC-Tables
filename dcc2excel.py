import openpyxl as pyxl
#import xml.etree.ElementTree as et
from lxml import etree as et
from DCChelpfunctions import search
from excel2dcc import printelement

LANG='da'

def write_row(write_sheet, row_num: int, starting_column: str or int, write_values: list):
    if isinstance(starting_column, str):
        starting_column = ord(starting_column.lower()) - 96
    for i, value in enumerate(write_values):
        write_sheet.cell(row_num, starting_column + i, value)

def statements(sheetname, statementselement):
    #WB.create_sheet(sheetname)
    line=1
    ws=WB[sheetname]
    columnheadings=['category','id', 'heading','body']
    write_row(ws,line,2,columnheadings)
    line+=1
    headingstr=DCC+"heading[@lang='"+LANG+"']"
    bodystr=DCC+"body[@lang='"+LANG+"']"
    for statement in statementselement.findall(DCC+'statement'):
        [category,ID,heading,body]=["","","",""]
        category=statement.find(DCC+'category').text
        ID=statement.attrib['statementId']
        heading=statement.find(headingstr).text
        body=statement.find(bodystr).text
        write_row(ws,line,2,[category,ID,heading,body])
        line+=1
def write_to_admin(ws, root, startline, section):
    line=startline
    for element in section.iter():
        head=[]
        for child in element.getchildren():
            if child.tag==DCC+"heading":
               head.append(child.text)
        if len(head):
            write_row(ws, line,1, head+ ['','', root.getpath(element)]) 
            line+=1
        if type(element.text)!=type(None): 
            if element.tag!=DCC+"heading": 
                if not(element.text.startswith('\n')):
                    write_row(ws, line,4,[element.text,root.getpath(element)]) 
                    line+=1

            
def admin(ws, root):
    toprow=["heading lang1", "heading lang2", "Description", "Value", "XPatht"]
    write_row(ws, 1 ,1, toprow) 
    head=[]
    for heading in root.findall(DCC+'heading'):
        head.append(heading.text)
    write_row(ws, 2 ,1, head) 
    root.find
    adm=root.find(DCC+"administrativeData")
    soft=adm.find(DCC+"dccSoftware")
    write_to_admin(ws ,root, 3, soft)
    core=adm.find(DCC+"coreData")
    write_to_admin(ws,root, 6, core)
    callab=adm.find(DCC+"calibrationLaboratory")
    write_to_admin(ws, root, 18,callab)
    cust=adm.find(DCC+"customer")
    write_to_admin(ws, root, 29,cust)

    



if __name__=="__main__":
    #################
    #first argument is dcc xml file
    #second argument is excel template to use
    import sys
    args=sys.argv[1:]
    print(len(args))
    if len(args)==0:
        xmlfile="Examples/DFM-T220000.xml"
    else:
        xmlfile=args[0]
    if len(args)==2:
        WB=pyxl.load_workbook(args[1])
    else:    
        WB=pyxl.Workbook()
        WB.create_sheet('Table')
        WB.create_sheet('statements')
        WB.create_sheet('AdministrativeData')
   
    DCC='{https://dfm.dk}'

    root=et.parse(xmlfile)
    
    
    attributes=[['scope'],['dataCategory'],['measurand'],['unit'],['metaDataCategory'],['humanHeading']]
    tab=root.find(DCC+'measurementResults').find(DCC+'measurementResult').find(DCC+'table')
    headingstr=DCC+"heading[@lang='"+LANG+"']"
    
    cols=[]
    for col in tab.findall(DCC+'column'):
        attributes[0].append(col.attrib['scope'])
        attributes[1].append(col.attrib['dataCategory'])
        attributes[2].append(col.attrib['measurand'])
        attributes[3].append(col.find(DCC+'unit').text)
        attributes[4].append(col.attrib['metaDataCategory'])
        attributes[5].append(col.find(headingstr).text)
        col=search(root, tab.attrib,col.attrib,col.find(DCC+'unit').text)[0][-1].text.split()
        if col=='-':
            col=['']*len(cols[0])
        cols.append(col)
    
    
    ws=WB['Table']
    line=1
    write_row(ws,line,1,["DCCTable"])
    line+=1
    for item in tab.attrib.items():
        write_row(ws,line,1,[item[0],item[1]])
        line+=1
    
    write_row(ws,line,1,['numRows',len(cols[0])])
    line+=1
    write_row(ws,line,1,['numColumns',len(cols)])
    line+=3
    
    for row in attributes:
        write_row(ws,line,1,row)
        line+=1
        #ws.append(row)
    for n in range(0,len(cols[0])):
        r=['']+[c[n] for c in cols]
        write_row(ws,line,1,r)
        line+=1
        #ws.append(r) 

    stat=root.find(DCC+'administrativeData').find(DCC+'statements')
    statements('statements',stat)
    admin(WB['AdministrativeData'],root)

    WB.save("view_content.xlsx")
         
