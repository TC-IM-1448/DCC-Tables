import openpyxl as pyxl
import xml.etree.ElementTree as et
from xml.dom import minidom
import DCChelpfunctions as DCCh
#Used from DCChelpfunctions :
#validate, DCC_tablecolumn

DCC='{https://dfm.dk}'
et.register_namespace("dcc", DCC.strip('{}'))
lang1='en'
lang2='da'
languages={'lang1':'en','lang2':'da'}

def dictionaries_from_table(ws, rowtype='x'):
    """
    Input: openpyxl worksheet object with column headers and row headers.
    Input: rowheader to look for (string)
    for each row of the given type make a dictionary with keys defined by the column headings and values defined in the cells of that row.
    Return the dictionaries as a list
    """
    columnheaders=ws['1']
    rowheaders=ws['A']
    dictionaries=[]
    rowno=0
    for rowheader in rowheaders:
        rowno+=1
        if rowheader.value==rowtype:
            dictionary={}
            for (name, content) in zip(columnheaders, ws[rowno]):
                dictionary[name.value]=content.value
            dictionaries.append(dictionary)
    return(dictionaries)

def read_statements_from_Excel(root, ws):
    adm=root.find(DCC+"administrativeData")
    statements=dictionaries_from_table(ws)
    statementselement=et.SubElement(adm,DCC+"statements")
    for statement in statements:
        statementelement=et.SubElement(statementselement,DCC+"statement", attrib={'statementId':statement['id']})
        et.SubElement(statementelement, DCC+"category").text=statement['category']
        et.SubElement(statementelement, DCC+"heading", attrib={'lang':lang1}).text=statement['heading lang1']
        et.SubElement(statementelement, DCC+"heading", attrib={'lang':lang2}).text=statement['heading lang2']
        et.SubElement(statementelement, DCC+"body", attrib={'lang':lang1}).text=statement['body lang1']
        et.SubElement(statementelement, DCC+"body", attrib={'lang':lang2}).text=statement['body lang2']
    return root

def read_accreditation_from_Excel(root, ws):
    acc=dictionaries_from_table(ws)
    adm=root.find(DCC+"administrativeData")
    accelement=et.SubElement(adm,DCC+"accreditation", attrib={'accrId':acc[0]['id']})
    for key, value in acc[0].items():
        if type(key)!=type(None) and "acc" in key:
            et.SubElement(accelement,DCC+key).text=str(value)
    return root

def read_settings_from_Excel(root, ws):
    settings=dictionaries_from_table(ws)
    adm=root.find(DCC+"administrativeData")
    settingselement=et.SubElement(adm,DCC+"settings")
    for setting in settings:
        settingelement=et.SubElement(settingselement,DCC+"setting", attrib={'settingId':setting['id']})
        et.SubElement(settingelement, DCC+"heading", attrib={'lang':lang1}).text=setting['heading lang1']
        et.SubElement(settingelement, DCC+"heading", attrib={'lang':lang2}).text=setting['heading lang2']
        et.SubElement(settingelement, DCC+"body", attrib={'lang':lang1}).text=setting['body lang1']
        et.SubElement(settingelement, DCC+"body", attrib={'lang':lang2}).text=setting['body lang2']
        if type(setting['value'])!=type(None):
           et.SubElement(settingelement, DCC+"value").text=str(setting['value'])
        if type(setting['unit'])!=type(None):
           et.SubElement(settingelement, DCC+"unit").text=setting['unit']
    return root

def read_equipment_from_Excel(root, ws):
    """
    Parameters
    ----------
    root : etree element
    ws : openpyxl worksheet object
        DESCRIPTION.

    Returns
    root element updated with items section
    -------
    None.
    """
    items=dictionaries_from_table(ws)

    #Make an item XML-element
    administrativeData=root.find(DCC+"administrativeData")
    Itemslist=et.SubElement(administrativeData,  DCC+'items')
    Itemslist=root.find(DCC+"administrativeData").find(DCC+"items")
    for item in items:
        equipmentelement=et.SubElement(Itemslist,DCC+"equipment",attrib={'equipId':item['id'], 'category':item['category']})
        for key,value in languages.items():
            et.SubElement(equipmentelement, DCC+"heading", attrib={'lang':value}).text=item['heading '+key]
        for key in ['manufacturer', 'productName', 'productNumber']:
            if type(item[key])!=type(None):
                et.SubElement(equipmentelement, DCC+key).text=item[key]
        for idn in ['id1','id2']:
     
            if type(item[idn])!=type(None):
                identelement=et.SubElement(equipmentelement, DCC+"identification", attrib={'issuer':item[idn+' issuer']})
            for key,value in languages.items():
                et.SubElement(identelement, DCC+"heading", attrib={'lang':value}).text=item[idn+' heading '+key]
            et.SubElement(identelement, DCC+"value").text=item[idn]
    return root

def read_admin_from_Excel(root, ws):
    headingslang1=ws['A'][1:]
    headingslang2=ws['B'][1:]
    DFM_names=ws['C'][1:]
    values=ws['D'][1:]
    xpaths=ws['E'][1:]
    for i, path in enumerate(xpaths):
        try:
            levels=path.value.split("/dcc:")[2:]
        except:
            levels=[]
        element=root
        for level in levels:
            subelement=element.find(DCC+level)
            if type(subelement)!=type(None):
                element=subelement
            else:
                element=et.SubElement(element,DCC+level)
        element.text=values[i].value
        if type(headingslang1[i].value)!=type(None):
            et.SubElement(element,DCC+"heading",attrib={'lang':lang1}).text=headingslang1[i].value
        if type(headingslang2[i].value)!=type(None):
            et.SubElement(element,DCC+"heading",attrib={'lang':lang2}).text=headingslang2[i].value
    return root

def read_table_from_Excel(root, ws, cell0):

    columns = []
    attrib={}

    headings={}
    headings['lang1']=cell0.offset(1,2).value
    headings['lang2']=cell0.offset(1,3).value
    attribnames=[cell0.offset(r,0) for r in range(1,5)]
    attribvalues=[cell0.offset(r,1) for r in range(1,5)]
    for (name, value) in zip(attribnames,attribvalues):
        if type(name.value) != type(None) and type(value.value) != type(None):
            attrib[name.value]=value.value
    numRows = cell0.offset(6,1).value
    numColumns = cell0.offset(7,1).value

    numHeadings=7
    nRows = int(numRows)+numHeadings
    nCols = int(numColumns)
    cell = cell0.offset(8,1)

    content = [[cell.offset(r,c).value for r in range(nRows)] for c in range(nCols)]

    for c in content:
        col = DCCh.DccTableColumn(   scopeType=c[0],
                                columnType=c[1],
                                measurandType=c[2],
                                unit=c[3],
                                metaDataCategory=c[4],
                                humanHeading = [c[5],c[6]],
                                columnData= list(map(str, c[numHeadings:])))
        columns.append(col)

    #Create empty table with table attributes
    xmltable1=et.Element(DCC+"table",attrib=attrib)
    for key,value in languages.items():
        et.SubElement(xmltable1, DCC+"heading", attrib={'lang':value}).text=headings[key]
    
    et.SubElement(xmltable1,DCC+"numrows").text=str(numRows)
    et.SubElement(xmltable1,DCC+"numcols").text=str(numColumns)

    #Fill the table with data from table object
    columns=columns
    for col in columns:
        attributes={'scope':col.scopeType, 'dataCategory':col.columnType, 'measurand':col.measurandType, 'metaDataCategory':col.metaDataCategory}
        xmlcol=et.Element(DCC+'column',attrib=attributes)
        for key,value in languages.items():
            et.SubElement(xmlcol, DCC+"heading", attrib={'lang':value}).text=col.humanHeading[key]
        if type(col.unit)!=type(None):
          et.SubElement(xmlcol,DCC+'unit').text=' '.join([col.unit])
        #DCCh.add_name(xmlcol,lang="en",text=col.humanHeading)
        #xmllist=realListXMLList(value=col.columnData,unit=[col.unit])
        if attributes['metaDataCategory']=='Data':
           if attributes['dataCategory']=='Conformity':
               et.SubElement(xmlcol,DCC+"conformityXMLList").text=' '.join(col.columnData)
           else:
               et.SubElement(xmlcol,DCC+"valueXMLList").text=' '.join(col.columnData)
        elif attributes['metaDataCategory']=='customerTag':
            et.SubElement(xmlcol,DCC+"stringXMLList").text=' '.join(col.columnData)
        elif attributes['metaDataCategory']=='Exception':
            et.SubElement(xmlcol,DCC+"exceptionXMLList").text=' '.join(col.columnData)
        elif attributes['metaDataCategory']=='accreditationException':
            et.SubElement(xmlcol,DCC+"exceptionXMLList").text=' '.join(col.columnData)
        else:
            et.SubElement(xmlcol,DCC+"stringXMLList").text=' '.join(col.columnData)
        xmltable1.append(xmlcol)
    measurementResults=root.find(DCC+"measurementResults")
    measurementResult=measurementResults.find(DCC+"measurementResult")
    if type(measurementResult)==type(None):
        measurementResult=et.SubElement(measurementResults,DCC+'measurementResult', attrib={'resId':'result1'})
    measurementResult.append(xmltable1)
    return root

def printelement(element):
    xmlstring=minidom.parseString(et.tostring(element)).toprettyxml(indent="   ")
    print(xmlstring)
    return
def minimal_DCC():
    version="1.0.0"
    xsilocation= DCC.strip('{}') + " dcc.xsd"
    xsi="http://www.w3.org/2001/XMLSchema-instance"
    root=et.Element(DCC+'digitalCalibrationCertificate',
            attrib={"schemaVersion":version,   "xmlns:xsi":xsi, "xsi:schemaLocation":xsilocation})
    return root

if __name__ == "__main__":
    import sys
    from importlib import reload
    reload(DCCh)
    args=sys.argv[1:]
    if len(args)==1:
        workbookName=args[0]
    else:
        workbookName="Examples/DCC_temperature.xlsx"
    schema      ="dcc.xsd"

    #load workbook
    wb = pyxl.load_workbook(workbookName, data_only=True)
    #Create root element with minimal content
    root = minimal_DCC()

    #Update root element with content from worksheets in the workbook
    root = read_admin_from_Excel(     root, ws=wb["AdministrativeData"])
    root = read_accreditation_from_Excel(root, ws=wb["Accreditation"])
    root = read_statements_from_Excel(root, ws=wb["Statements"])
    root = read_equipment_from_Excel(      root, ws=wb["Equipment"])
    root = read_settings_from_Excel(root, ws=wb["Settings"])
    measurementResults=et.SubElement(root,DCC+'measurementResults')
    #All sheets whose name contain "Table" will be interpreted as table-sheets
    tableSheets=[]
    for sheetname in wb.sheetnames:
        if "Table" in sheetname:
            tableSheets.append(sheetname)
    for tableSheet in tableSheets:
        for cell in wb[tableSheet]['A']:
            if cell.value=='DCCTable':
               root = read_table_from_Excel(root, ws=wb[tableSheet], cell0=cell)
    wb.close()

    ############### Output to xml-file ####################################
    outputxml = root.find(DCC+"administrativeData").find(DCC+"coreData").find(DCC+"uniqueIdentifier").find(DCC+'value').text+".xml"
    xmlstr=minidom.parseString(et.tostring(root)).toprettyxml(indent="   ")
    with open(outputxml,'wb') as f:
        f.write(xmlstr.encode('utf-8'))

    #Validate the ouptut file against the schema and output the result.
    print(DCCh.validate(outputxml, schema))
