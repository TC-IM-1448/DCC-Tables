import os
import openpyxl as pyxl
import xml.etree.ElementTree as etree
import DCChelpfunctions as dcchf
from DCChelpfunctions import search

LANG='da'
DCC='{https://dfm.dk}'

def lookupFromMappingFile(mapFileName:str, dccFileName:str):
    """LookupFromMappingFile """
    root = etree.parse(dccFileName)

    wb = pyxl.load_workbook(mapFileName)
    sheet = wb['Mapping']
    n_rows = sheet.max_row
    n_cols = sheet.max_column

    for i in range(2, n_rows-1):
        # Itterate over the first column in the mapping sheet until '--END--' is reached or last row
        cellA = sheet.cell(row=i, column=1).value
        if cellA == "--END--": break
        
        queryType = str(sheet.cell(row=i, column=3).value)
        # print(queryType, end = "   ")

        # distinguish between two query types: xpath and data.
        if queryType == 'xpath': 
            cellC = sheet.cell(row=i, column=4).value
            xpath = cellC
            if xpath.startswith("/dcc:digitalCalibrationCertificate"):
                s = './'+xpath.split("/dcc:digitalCalibrationCertificate")[1]
            else: 
                s = '.'+xpath
            ss = s.replace("dcc:", DCC)
            # print(ss)
            elm = root.findall(ss)[0]
            elm = elm.text
            cell = sheet.cell(row=i, column=n_cols+1, value=elm)
        elif queryType == 'data':
            tableId = sheet.cell(row=i, column=5).value
            itemRef = sheet.cell(row=i, column=6).value
            settingRef = sheet.cell(row=i, column=7).value
            scope = sheet.cell(row=i, column=8).value
            dataCategory = sheet.cell(row=i, column=9).value
            measurand = sheet.cell(row=i, column=10).value
            metaDataCategory = sheet.cell(row=i, column=11).value
            unit = sheet.cell(row=i, column=12).value
            customerTag = sheet.cell(row=i, column=13).value
            data = search(root,
                          {'tableId': tableId, 
                            'itemRef': itemRef, 
                            'settingRef': settingRef, 
                            }, 
                           {'scope': scope, 
                            'dataCategory': dataCategory, 
                            'measurand': measurand, 
                            'metaDataCategory': metaDataCategory
                            }, 
                            unit = unit, 
                            customerTag = customerTag
                            )[0]
            cell = sheet.cell(row=i, column=n_cols+1, value=data)
        else: 
            cell = sheet.cell(row=i, column=n_cols+1, value="FAILED")


    outFileName = mapFileName.rsplit(".",maxsplit=1)[0]+'_QueryResult.xlsx'
    wb.save(outFileName)

if __name__=="__main__":
    #################
    #first argument is dcc xml file
    #second argument is excel template to use

    import sys
    args=sys.argv[1:]
    print(len(args))
    if len(args)==0:
        mapFileName ='Examples'+os.sep+'Mapping_Novo_temperatur_Certifikat.xlsx'
        dccFileName = 'Examples'+os.sep+'Stip-230063-V1.xml'
        lookupFromMappingFile(mapFileName, dccFileName)
    elif len(args)==2:
        mapFileName = args[0]
        dccFileName = args[1]
        lookupFromMappingFile(mapFileName, dccFileName)
    else: 
        helpstatement = """call dccquery.py using the following arguments: \n 
        >> python dccquery.py [mapping file e.g. mapping.xlsx] [DCC file e.g. dcc.xml] """
        print(helpstatement)

         
