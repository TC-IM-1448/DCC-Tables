# 
# by DBH 2024-01-11
#
# Uses xlwings see following references: 
# https://docs.xlwings.org/en/stable/syntax_overview.html 
# https://docs.xlwings.org/en/latest/ 

__ver__ = "v0.0.3"

import os
import re
import base64
import openpyxl as pyxl
import xlwings as xw
from  lxml import etree as et
from lxml import builder as etb
from lxml.builder import ElementMaker 
import tkinter as tk
from tkinter import ttk
import tkinter.filedialog as tkfd
import DCChelpfunctions as dcchf
from DCChelpfunctions import search


LANG='da'
NUM_LANGS = 0
DCC='{https://dfm.dk}'

xlValidateList = xw.constants.DVType.xlValidateList


SECTION_HEADINGS = {'statementList':[],
                    'equipmentList':[],
                    'settingList':[],
                    'measurementConfigList':[],
                    'quantityUnitDefList':[],
                    'embeddedFileList':[],
                    'table':[],
                    'column':[],
                    'administrativeData':[],
                    'contactAndLocation':[],
                    'address':[],
                    'contactInfo':[],
                    'geoPosition':[],
                    }

HEADINGS = SECTION_HEADINGS.copy()

# dict(statementList = ['in DCC', 
#                                     '@id', 
#                                     '@category', 
#                                     '@imageRefs',
#                                     'heading[1]',  
#                                     'body[1]', 
#                                     'externalReference'],

#     equipmentList = ['in DCC', '@id', '@statementRef', '@category',
#                         'heading[1]', 'manufacturer', 'modelName', 'modelNumber', 
#                         'serialNumber', 'lotNumber', 'productClass',
#                         'clientId', 'serviceProviderId',
#                         'prevCalibDate','calibDueDate', 'prevCertId', 'prevCertProviderName',
#                         ],

#     settingList = [ 'in DCC', '@id', '@equipmentRef', 
#                     'parameter', 'value', 'unit', 'softwareInstruction', 
#                     'heading[1]', 'body[1]', #'statementRefs',
#                     ],

#     measurementConfigList = [   'in DCC', '@id', 
#                                 'heading[1]', 
#                                 'equipmentRefs', 'settingRefs', 'statementRefs',
#                                 'operationalStatus',
#                                 'body[1]', 
#                                 ], 

#     embeddedFileList = ['in DCC', '@id', 
#                         'heading[1]',
#                         'body[1]',
#                         'fileExtension'],

#     quantityUnitDefList = [ 'in DCC', 
#                             '@id',
#                             '@quantityCodeSystem',
#                             'quantityCode',
#                             'unitUsed',
#                             'functionToSIunit',
#                             'unitSI',
#                             'externalRefs',
#                             'heading[1]',
#                             '@statementRefs'],

#     administrativeData = [  "level", 
#                             "description",
#                             "@value", 
#                             "heading[1]",  
#                             "xsdType",
#                             "xPath"], 

#     table = [   'tableCategory', 
#                 '@tableId', 
#                 '@serviceCategory', 
#                 '@measurementConfigRef', 
#                 '@customServiceCategory', 
#                 '@statementRefs',
#                 '@embeddedFileRefs',
#                 'heading[1]', 
#                 'conformityStatus', 
#                 'conformityStatusRef',
#                 '@numRows', 
#                 '@numCols'], 

#     column = [  'scope', 
#                 'dataCategory', 
#                 'dataCategoryRef', 
#                 'quantity', 
#                 'unit', 
#                 'quantityUnitDefRef', 
#                 'equipmentRef', 
#                 'heading[1]', 
#                 'idx']
#     )


class DccGuiTool(): 
    xsdDefInitCol = 3  
    xsdRestrictions = {}
    xsdRoot = None
    xsdTree = None
    dccRoot = None  
    dccTree = None  
    wb = None   
    sheetDef = None 
    embeddedFilesPath = None
    langs = []
    
    colors = dict(  white = "#ffffff",
                    black = "#000000",
                    yellow = "#ffd966",
                    light_yellow = '#fff2cc',
                    green = '#c6e0b4',
                    light_green = '#e2efda',
                    gray = '#bfbfbf',
                    light_gray = '#d9d9d9',
                    blue = '#bdd7ee',
                    light_blue = '#ddebf7',
                    red = '#f8cbad',
                    light_red = '#fce4d6')


    def loadExcelWorkbook(self, workBookFilePath: str):
        self.wb = xw.Book(workBookFilePath)
        wb = self.wb
        if not "Definitions" in wb.sheet_names:
            wb.sheets.add("Definitions")
        self.sheetDef = wb.sheets['Definitions']
        self.wb = wb
        self.loadSchemaFile()
        self.loadSchemaRestrictions()
        self.langs = ['en']
        wb.activate(steal_focus=True)
        self.createEmbeddedFilesFolder(wb)


    def createEmbeddedFilesFolder(self, wb):
        wbpath = wb.fullname.replace(wb.name, '')
        if not os.path.exists(wbpath+'embeddedFiles'):
            os.mkdir(wbpath+'embeddedFiles')
        self.embeddedFilesPath = wbpath+'embeddedFiles'+os.sep


    def loadSchemaFile(self, xsdFileName="dcx.xsd"):
        self.xsdTree, self.xsdRoot  = dcchf.load_xml(xsdFileName)
        

    def loadDCCFile(self, xmlFileName="dcc-example.xml"):
        self.dccTree, self.dccRoot = dcchf.load_xml(xmlFileName)
        self.schemaVersion = self.dccRoot.attrib['schemaVersion']
        errors = dcchf.validate(xmlFileName, 'dcx.xsd')    
        langs = dcchf.get_languages(self.dccRoot)

        for sht in self.wb.sheets: 
            if not sht.name == "Definitions": sht.delete()

        # stoere Metadata in the definitions sheet. 
        sht = self.wb.sheets[0]
        prog = os.path.basename(__file__)
        sht.range((2,1)). value = [['Program File',prog],
                                   ['Program Version', __ver__], 
                                   ['SchemaFile', 'dcx.xsd'],
                                   ['SchemaVersion', self.schemaVersion],
                                   ['xml file name', os.path.basename(xmlFileName)],
                                   ['xml file path', xmlFileName],
                                   ['xml languages', " ".join(langs)]]

        # continue loading data
        self.loadSchemaRestrictions()

        # Select languages
        self.chooseLanguages(langs + ['---']+self.xsdRestrictions['stringISO639Type'])
            # note chooseLanguages stores the selected languages in self.langs 
        sht.range((9,1)).value =['selected Languages', " ".join(self.langs)]
        self.selectedLangsRowIndex = 9

        # laod headings for tables from the xsd-schema. 
        self.loadXSDTableHeadings()

        # load the data into the tables. 
        self.loadDccSequence()
        return errors


    def loadSchemaRestrictions(self): 
        xsd_root = self.xsdRoot
        sht_def = self.sheetDef
        drestr = dcchf.schema_get_restrictions(xsd_root)
        self.wb.app.screen_updating = False
        rng = sht_def.range((1,3)).expand()
        rng.clear()
        j = self.xsdDefInitCol
        for i, (k,vs) in enumerate(drestr.items()):
            rng = sht_def.range((1, i+j))
            rng.value = [k]
            rng.offset(1,0).value = [[v] for v in vs]
            rng.offset(1,0).expand('down').name = k  
            rng.font.bold = True
        self.dccDefInitCol = i+1
        self.xsdRestrictions = drestr
        self.wb.app.screen_updating = True
        return drestr


    def chooseLanguages(self, languages):
        global app
        languageDialog = MyLanguageDialog(app, languages)
        app.wait_window(languageDialog.top)

    # def updateHeadingLanguages(self, langs):
    #     """ DBH: Probably obsolete function now. """
    #     global HEADINGS
    #     headings = {}
    #     for k,v in HEADINGS.items():
    #         new_v = v[:]
    #         for l in ['heading[1]', 'body[1]']:
    #             hidxs = [i for i,s in enumerate(new_v) if l in s]
    #             for i in hidxs[::-1]:
    #                 h = new_v[i] 
    #                 new_h = [h.replace('1', lang) for lang in langs]
    #                 new_v[i:i+1] = new_h
    #         headings[k] = new_v
    #     self.headings = headings
    #     return self.headings

    def resizeXlTable(self,rng,sht,tableName:str):
        if tableName not in [tbl.name for tbl in sht.tables]:
            sht.tables.add(source=rng, name=tableName)
        else:
            sht.tables[tableName].resize(rng)

    def getHeadingOrBodyFromXlHeadingTag(self, node: et._Element, headingTag:str) -> str: 
        """Get the heading or body text from the node based on the headingTag"""
        h = headingTag

        lang = h[h.index('[')+1:h.index(']')]
        headOrBody = h.split('[')[0]
        searchStr = f'./dcx:{headOrBody}[@lang="{lang}"]'
        nodes = node.findall(searchStr, node.nsmap)
        if len(nodes) > 0: return nodes[0].text
        else: return None

    def loadDccSequence(self):
        self.loadDCCAdministrativeInformation(after='Definitions',
                                              heading=self.headings['administrativeData']
                                              )
        
        self.loadDccInfoTable(heading = self.headings['statementList'], 
                                nodeTag="dcx:statementList",
                                subNodeTag="dcx:statement",
                                place_sheet_after='AdministrativeData')
        
        self.loadDccInfoTable(heading = self.headings['equipmentList'], 
                                nodeTag="dcx:equipmentList", 
                                subNodeTag="dcx:equipment",
                                place_sheet_after='statementList')
        
        self.loadDccInfoTable( heading = self.headings['settingList'], 
                                nodeTag="dcx:settingList", 
                                subNodeTag="dcx:setting",
                                place_sheet_after='equipmentList')
        
        self.loadDccInfoTable( heading = self.headings['measurementConfigList'], 
                                nodeTag="dcx:measurementConfigList", 
                                subNodeTag="dcx:measurementConfig",
                                place_sheet_after='settingList')
        
        self.loadDccInfoTable( heading = self.headings['embeddedFileList'], 
                                nodeTag="dcx:embeddedFileList", 
                                subNodeTag="dcx:embeddedFile",
                                place_sheet_after='measurementConfigList')
        
        self.loadDccInfoTable(heading=self.headings['quantityUnitDefList'],
                              nodeTag="dcx:quantityUnitDefList", 
                              subNodeTag="dcx:quantityUnitDef",
                              place_sheet_after="embeddedFileList")
        
        self.loadDCCMeasurementResults(heading=self.headings['table'])


    def loadXSDTableHeadings(self):
        xsd_root = self.xsdRoot
        langs = self.langs
       
        #%% 
        global SECTION_HEADINGS
        headings = SECTION_HEADINGS.copy()
        for secTag in headings.keys():
            # build infoTable heading from schema dcx.xsd.
            secTypeTag = secTag.replace("List","") + "Type"
            tblHeading, tblHeadingType = dcchf.getNamesAndTypes(xsd_root, secTypeTag, exclude_heading=False)
            if secTag == "administrativeData":
                 tblHeading = [ "level", 
                                "description",
                                "@value",
                                "heading",
                                "xsdType",
                                "xPath"]
            # insert selected languages
            for k in ['heading', 'body']:
                if k in tblHeading:
                    i = tblHeading.index(k)
                    tblHeading[i:i+1] = [f"{k}[{l}]" for l in langs]
            # insert column for selecting rows to export. 
            if "List" in secTag:
                tblHeading[0:0] = ['in xml']
            
            headings[secTag] = tblHeading

        # Customize particular headings. 
        headings['column'][1:1]=['dataCategory']
        i = headings['table'].index('column')
        headings['table'].pop(i)
        headings['table'].insert(0,'tableCategory')
        headings['embeddedFileList'].pop(-1)
        
        global HEADINGS
        HEADINGS = headings.copy()
        #%%
        self.headings = headings


        # print the constructed dictionary 
        for H in headings.keys(): 
            print(f"{H}:") 
            for h in headings[H]: print(f"\t{h}")
        
        #%%

    def loadDccInfoTable(self, 
                         heading=['in DCC', '@category', '@statementId', 
                                  'heading[1]', 'body[1]', 
                                  'heading[2]', 'body[2]'], 
                         nodeTag="dcx:statements", 
                         subNodeTag="dcx:statement",
                         place_sheet_after='AdministrativeData' ): 
        root = self.dccRoot
        ns = root.nsmap
        wb = self.wb
        langs = self.langs
        N = len(self.langs)


        # insert the sheet if it is missing
        shtName = nodeTag.replace('dcx:','')    # statementList, equipmentList, ...
        # subNodeAttrId = shtName[:-1]+'Id'        # statementId
        if not shtName in wb.sheet_names:
            wb.sheets.add(shtName, after=place_sheet_after)
        sht = wb.sheets[shtName]
        node = root.find(".//"+nodeTag,ns) 

        # Write the Sheet headings
        sht.range((1,1)).value = [[f'heading[{lang}]'] for lang in langs]
        nodeHeadings = node.findall("dcx:heading",ns)
        for h in nodeHeadings:
            lang = h.attrib['lang']
            if lang in langs: 
                sht.range((langs.index(lang)+1,2)).value = h.text
    
        rng = sht.range((1,1),(N,1))
        rng.api.Borders.Weight = 2
        rng.color = self.colors['gray']
        rng = sht.range((1,2),(N,2))
        rng.api.Borders.Weight = 2
        rng.color = self.colors['blue']
        rng.font.bold = True



        #%%
            
        # Write the table column headings and set column widths
        
        tblRowIdx = len(langs)*2+1
        sht.range((tblRowIdx,1)).value = heading
        sht.range((1,1),(N,len(heading))).columns.autofit()

        hIdxs = [idx for idx,val in enumerate(heading) if val.startswith('heading')]
        bIdxs = [idx for idx,val in enumerate(heading) if val.startswith('body')]
        for i in hIdxs:
            sht.range((1,i+1)).column_width = 30
        for i in bIdxs:
            sht.range((1,i+1)).column_width = 50

        # Write human-readable column headings. 
        #%%
        if "List" in shtName: 
            headings = self.headings
            langs = self.langs
            numLangs = len(langs)
            tblHeadings = headings[shtName]
            n = len(tblHeadings)
            humanHeadings = [[None]*n for i in range(numLangs)]
            node = root.find(".//dcx:"+shtName,root.nsmap)
            nodeColHead =  dcchf.xpath_query(node, "./dcx:columnHeadings")
    
            if len(nodeColHead):
                nodeColHead = nodeColHead[0]
                for j,l in enumerate(langs):
                    for i,h in enumerate(tblHeadings):
                        hNoLang = removeHeadingLang(h)
                        n = dcchf.xpath_query(nodeColHead, f".//dcx:column[@name='{hNoLang}']/dcx:heading[@lang='{l}']")
                        if len(n) > 0:
                            txt = n[0].text
                            humanHeadings[j][i] = txt
                        else:
                            txt = None
                            humanHeadings[j][i] = txt
                        print(i,j,l,h,txt)
                print(shtName, langs)
                print(humanHeadings)
            
            sht.range((N+1,1)).value = humanHeadings # column index is 1, because "in xml" is the first column in tblHeadings
            rng = sht.range((N+1,2),(N*2,len(tblHeadings)))
            rng.api.Borders.Weight = 2
            rng.color = self.colors['light_blue']
            
            rng = sht.range((N+1,1),(N*2,1))
            rng.api.Borders.Weight = 2
            rng.color = self.colors['light_gray']
            rng.value = [[f"colHead[{l}]"] for l in langs]

        #%%

        # load the information into the table
        tableData = []
        rows = node.findall(subNodeTag,ns)
        for idx, subNode in enumerate(rows):
            rowData = []
            for i, h in enumerate(heading):
                if i == 0: 
                    rowData.append('y')
                    # if is an attribute
                elif h.startswith('@'): 
                    if h.strip('@') in subNode.attrib.keys():
                        rowData.append(subNode.attrib[h.strip('@')])
                    else: 
                        rowData.append(None)
                elif h.startswith('heading['): 
                    rowData.append(self.getHeadingOrBodyFromXlHeadingTag(subNode, h))
                elif h.startswith('body['):
                    rowData.append(self.getHeadingOrBodyFromXlHeadingTag(subNode, h))
                elif nodeTag == "dcx:quantityUnitDefList" and h.find("unit") >= 0: 
                    nodes =  subNode.findall(f'./dcx:{h}', ns)
                    if len(nodes)>0: rowData.append("'"+nodes[0].text)
                    else: rowData.append(None)
                else: 
                    nodes =  subNode.findall(f'./dcx:{h}', ns)
                    if len(nodes)>0: rowData.append(nodes[0].text)
                    else: rowData.append(None)
                if nodeTag == "dcx:embeddedFileList" and h=="@id": 
                    # save the embedded file to temporary folder. 
                    fileId = subNode.attrib[h.strip('@')]
                    filePath = self.embeddedFilesPath+fileId
                    fileData = subNode.find("dcx:fileContent", ns).text
                    fileIsSaved = self.saveEmbeddedFileToFolder(filePath, fileData)
                    cidx = len(heading)+1
                    # show the file in the sheet if the files is an image. 
                    if fileId.split('.')[-1].lower() in ['png', 'emf', 'jpg'] and fileIsSaved:
                        sht.pictures.add(filePath, name=fileId, anchor=sht.range((4+idx,1+cidx+idx)))
                    sht.range((tblRowIdx+1+idx,cidx)).value = filePath
                    sht.range((1,2)).column_width = 27

            tableData.append(rowData)
            # rng = sht.range((tblRowIdx+1+idx,1))
            # rng.value = rowData

            if nodeTag == "dcx:measuringSystems": 
                sht.range((1,3)).column_width = 30


        rng = sht.range((tblRowIdx+1,1))
        rng.value = tableData
        rng = sht.range((tblRowIdx,1),(tblRowIdx+len(rows),len(heading)))
        self.resizeXlTable(rng,sht,'Table_'+shtName)
        rng.api.WrapText = True
        rng.columns

        if shtName == "statementList": 
            #Apply statement category validator to the statement@category column
            rng = sht.range("Table_"+shtName+"['@category]") 
            self.applyValidationToRange(rng, 'statementCategoryType')
            statementIdRng = wb.sheets[shtName].range("Table_"+shtName+"['@id]")
            statementIdRng.name = "statementListIdRange"
        
        if shtName == "equipmentList":
            #Apply equipment category type validator to the equipment@category column
            rng = sht.range("Table_"+shtName+"['@category]")
            self.applyValidationToRange(rng, 'equipmentCategoryType')
            # Give a name to the equipmentId column
            equipIdRng = wb.sheets[shtName].range("Table_"+shtName+"['@id]")
            equipIdRng.name = "equipmentListIdRange"
        
        if shtName == "settingList":
            #Apply equipmentId validator to the setting@refId column
            rng = sht.range("Table_"+shtName+"['@equipmentRef]")
            self.applyValidationToRange(rng, 'equipmentListIdRange')
            settingIdRng = wb.sheets[shtName].range("Table_"+shtName+"['@id]")
            settingIdRng.name = "settingListIdRange"

        if shtName == 'measurementConfigList': 
            # Give a name to the measurementId column
            measuringSysIdRng = sht.range("Table_"+shtName+"['@id]")
            measuringSysIdRng.name = "measurementConfigIdRange"
            #Apply operationalStatus validator to the measuringSystems@operationalStatus column
            rng = sht.range("Table_"+shtName+"[operationalStatus]")
            self.applyValidationToRange(rng, 'operationalStatusType')

        if shtName == "quantityUnitDefList":
            #Apply equipmentId validator to the setting@refId column
            rng = sht.range("Table_"+shtName+"['quantityCodeSystem]")
            self.applyValidationToRange(rng, 'quantityCodeSystemType')
            quIdRng = sht.range("Table_"+shtName+"['@id]") 
            quIdRng.name = "quantityUnitDefListIdRange"

        if shtName == 'embeddedFileList': 
            # Give a name to the measurementId column
            measuringSysIdRng = sht.range("Table_"+shtName+"['@id]")
            measuringSysIdRng.name = "embeddedFilesIdRange"
            rng = wb.sheets['statementList'].range("Table_statementList['@imageRefs]")
            self.applyValidationToRange(rng, 'embeddedFilesIdRange')
        # Apply Validation
        # validatorMap= {'dcx:statements': }

    def applyValidationToRange(self, rng, restrictionName:str):
            # insert validation on table headings
            formula = '='+restrictionName
            rng.api.Validation.Delete()
            rng.api.Validation.Add(Type=xlValidateList, Formula1=formula) 

    def saveEmbeddedFileToFolder(self, filepath, base64_string):
        try:
            image_bytes = base64.b64decode(base64_string)

            with open(filepath, "wb") as img_file:
                img_file.write(image_bytes)
                
            print(f"Image saved to {filepath}")
            return True
        except Exception as e:
            print(f"Error decoding base64 string: {e}")
            return False
        
    def loadDCCMeasurementResults(self, heading=['tableId', 
                                                 'tableCategory', 
                                                 'serviceCategory', 
                                                 'measuringSystemRef', 
                                                 'customServiceCategory', 
                                                 'statementRef',
                                                 'Heading Lang1', 
                                                 'Heading Lang2', 
                                                 'numRows', 
                                                 'numCols']):
        root = self.dccRoot
        ns = root.nsmap
        wb = self.wb

        measurementResults = root.find(".//dcx:measurementResultList",ns) 
        print(measurementResults)
        idTag = 'tableId'
        resultId = [c.attrib[idTag] for c in measurementResults.getchildren() if idTag in c.attrib]
        resultNodes = [c for c in measurementResults.getchildren() if idTag in c.attrib]
        calibrationResults = measurementResults.findall("dcx:calibrationResult",ns)
        calibResIds = [tbl.attrib[idTag] for tbl in calibrationResults]
        measurementSeries = measurementResults.findall("dcx:measurementSeries",ns)
        measSerIds = [tbl.attrib[idTag] for tbl in measurementSeries]

        for tableId in resultId: 
            if not tableId in wb.sheet_names:
                wb.sheets.add(tableId, after=wb.sheet_names[-1])
        sht = wb.sheets[resultId[0]]
        sht.activate()

        # dbh3
        for tbl in resultNodes: 
            tblType = dcchf.rev_ns_tag(tbl).split(':')[-1]
            tableId = tbl.attrib[idTag]
            # print(f"{tblId} : {tblCategoryType}")
            tabelHeadings = heading
                        
            sht = wb.sheets[tableId]
            sht.range("A1").value = [[txt] for txt in tabelHeadings]
            sht.range("A1").expand('down').columns.autofit()
            

            for i,h in enumerate(tabelHeadings):
                idx = i+1
                if h=='tableCategory':
                    # idx = tabelHeadings.index('tableCategory')+1
                    rng = sht.range((idx,2))
                    rng.value = tblType
                    self.applyValidationToRange(rng, 'tableCategoryType')
                # elif h == '@serviceCategory':
                #     rng = sht.range((idx,2))
                #     try:
                #         rng.value = tbl.attrib['serviceCategory']
                #     except KeyError:
                #         rng.value = None
                #     self.applyValidationToRange(rng,'serviceCategoryType')
                elif h == '@measurementConfigRef': 
                    rng = sht.range((idx,2))
                    rng.value = tbl.attrib['measurementConfigRef']
                    self.applyValidationToRange(rng, 'measurementConfigIdRange')
                elif h.startswith('heading['): 
                    rng = sht.range((idx,2))
                    rng.value = self.getHeadingOrBodyFromXlHeadingTag(tbl, h)
                elif h.startswith('@'): 
                    rng = sht.range((idx,2))
                    if h.strip('@') in tbl.attrib.keys():
                        rng.value = tbl.attrib[h.strip('@')]
                else:
                    rng = sht.range((idx,2)) 
                    nodes =  tbl.findall(f'./dcx:{h}', ns)
                    if len(nodes)>0: 
                        rng.value = nodes[0].text
                    else:
                        rng.value = None
                    if h+"Type" in dcchf.XSD_RESTRICTION_NAMES: 
                        self.applyValidationToRange(rng, h+'Type')
                    elif h in "conformityStatusRef": 
                        self.applyValidationToRange(rng,'statementListIdRange' )

            rng = sht.range((1,2), (idx,2))
            rng.color = self.colors["light_yellow"]
            rng.api.Borders.Weight = 2 
            
            colInitRowIdx = idx+2
            numRows = int(tbl.attrib['numRows'])
            numCols = int(tbl.attrib['numCols'])

            # Now load the columns
            #----------------------------------
            columns = tbl.findall("dcx:column", ns)
            # columnHeading = ['scope', 'dataCategory', 'dataCategoryRef', 'quantity', 'unit', 'quantityUnitDefRef', 'heading[1]', 'heading[2]', 'idx']
            columnHeading = self.headings['column']
            print(columnHeading)
            headingColorDefs = {'@scope': 'yellow',
                                 'dataCategory': 'yellow', 
                                 '@quantity': 'green',
                                 '@unit': 'green', 
                                 '@dataCategoryRef': 'light_yellow',
                                 '@quantityUnitDefRef': 'light_green', 
                                 '@settingRef': 'blue', 
                                 'heading': 'light_blue', 
                                 'idx': 'light_gray'
                                 }
            # headingColors = ['yellow', 'yellow', 'light_yellow', 'green', 'green', 'light_green', 'light_blue', 'light_blue', 'light_gray']
            headingColors = [self.colors[headingColorDefs[k.split('[', 1)[0]]] for k in columnHeading]

            sht.range((colInitRowIdx,1)).value = [[h] for h in columnHeading]
            # insert the index column
            rng = sht.range((colInitRowIdx+len(columnHeading),1))
            rng.value = [[i+1] for i in range(numRows)]
            rng = rng.expand('down')
            rng.color = headingColors[-1] 
            rng.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignCenter

            # 
            for colIdx, col in enumerate(columns):
                # insert the metadata heading attribute values
                cIdx = colIdx + 2
                for k, a in col.attrib.items():
                    rowIdx = colInitRowIdx + columnHeading.index("@"+k)
                    sht.range((rowIdx,cIdx)).value = "'"+a 

                # insert the dataCategory. 
                # dataList = col.find('dcx:dataList',ns)
                rowIdx = colInitRowIdx + columnHeading.index('dataCategory')
                dataCategory = dcchf.rev_ns_tag(col.getchildren()[-1])
                dataCategory = dataCategory.replace("dcx:", "", 1)
                sht.range((rowIdx,cIdx)).value = dataCategory 

                # instert the unit
                # unit = col.find("dcx:unit",ns).text
                # rowIdx = colInitRowIdx + columnHeading.index('unit')
                # sht.range(((rowIdx,cIdx))).value = unit                

                # Insert human readable heading in two languages. 
                for idx, lang in enumerate(self.langs):
                    xpath =  './/dcx:heading[@lang="{lang}"]'.format(lang=lang)
                    elm = col.find(xpath,ns)
                    if elm is None: 
                        val = None
                    else:
                        val = elm.text
                    rowIdx = colInitRowIdx + columnHeading.index(f'heading[{lang}]')
                    sht.range(((rowIdx,cIdx))).value = val


                # Insert the data 
                rowIdx = colInitRowIdx + len(columnHeading) - 1
                dataList = col.getchildren()[-1]
                dataPoints = {int(pt.attrib['idx']): pt.text for pt in dataList}
                if len(dataPoints) > 0:  
                    # dataType = dcchf.rev_ns_tag(dataList.getchildren()[0]).strip("dcx:")
                    # sht.range((rowIdx,cIdx)).value = dataType
                    for k,v in dataPoints.items(): 
                        sht.range((rowIdx+k,cIdx)).value = v 

            # set colors of the column heading rows
            for i in range(len(columnHeading)):
                idx = colInitRowIdx + i
                rng = sht.range((idx,1),(idx,numCols+1)) #).expand('right')
                rng.color = headingColors[i]
                
            # set Validator for the first four rows in the column heading-rows
            for i in [0,1,2,3]:
                idx = colInitRowIdx + i
                rng = sht.range((idx,2)).expand('right')
                rng.api.Validation.Delete()
                formula = columnHeading[i].strip('@') if not columnHeading[i][-3:] == "Ref" else 'dataCategory'
                formula = '='+formula+'Type'
                rng.api.Validation.Add(Type=xlValidateList, Formula1=formula) 

            rng = sht.range((colInitRowIdx+5,2),(colInitRowIdx+5,numCols+2))
            rng.api.Validation.Delete()
            rng.api.Validation.Add(Type=xlValidateList, Formula1="=quantityUnitDefListIdRange")

            rng = sht.range((colInitRowIdx+6,2),(colInitRowIdx+6,numCols+2))
            rng.api.Validation.Delete()
            rng.api.Validation.Add(Type=xlValidateList, Formula1="=settingListIdRange")

            
            # Set the column widths
            idx = colInitRowIdx + len(columnHeading) - 1
            rng = sht.range((colInitRowIdx,1),(colInitRowIdx+4,1)).expand('right')
            rng.columns.autofit()
            rng = sht.range((1,1)).expand()
            rng.columns.autofit()
            # Set the borders visible for the table
            rng = sht.range((colInitRowIdx,1)).expand()
            rng.api.Borders.Weight = 2 
            # Set the color of the data range in the table. 
            rng = sht.range((colInitRowIdx+len(columnHeading),2),(colInitRowIdx+len(columnHeading)+numRows-1,1+numCols))
            rng.color = self.colors["light_red"]


    def write_to_admin(self, sht, root, startline, section) -> int:
        N = numLangs = len(self.langs)
        line=startline
        for element in section.iter():
            head=[]
            # for child in element.getchildren():
            #     if dcchf.rev_ns_tag(child) == "dcx:heading":
            #         head.append(child.text)

            head = [child.text for child in element.getchildren() if dcchf.rev_ns_tag(child) == "dcx:heading"]
            elmHeadings = {h.attrib['lang']: h.text for h in element.findall("dcx:heading",root.nsmap)}
            head = [elmHeadings[lang] if lang in elmHeadings.keys() else '' for lang in self.langs ]
            if any(head):
                sht.range((line,1)).value = head + ['', '', self.dccTree.getpath(element)]
                sht.range((line,1),(line,N)).color = self.colors["light_yellow"]
                sht.range((line,N+1)).value = dcchf.rev_ns_tag(element).replace('dcx:','')
                sht.range((line,N+1)).font.bold = True
                line+=1
                for k,v in element.attrib.items(): 
                    sht.range((line,N+1)).value = "@"+k
                    rng = sht.range((line,N+2))
                    rng.value = str(v)
                    rng.color = self.colors["light_yellow"]
                    sht.range((line,N+3)).value = self.dccTree.getpath(element)
                    line+=1
            if type(element.text)!=type(None): 
                if dcchf.rev_ns_tag(element)!="dcx:heading": 
                    if not(element.text.startswith('\n')):
                        # print(element)
                        sht.range((line,N+1)).value = dcchf.rev_ns_tag(element).replace("dcx:",'')
                        sht.range((line,N+3)).value = self.dccTree.getpath(element)
                        rng = sht.range((line,N+2))
                        rng.value = str(element.text)
                        rng.color = self.colors["light_yellow"]
                        line+=1
        return line

             
    def loadDCCAdministrativeInformation(self, after='Definitions', 
                                         heading=[
                                                  "level",
                                                  "description", 
                                                  "@value", 
                                                  "heading[1]", 
                                                  "heading[2]", 
                                                  "xPath"]):
        root = self.dccRoot
        xsd_root = self.xsdRoot
        ns = root.nsmap
        wb = self.wb
        colors = self.colors
        langs = self.langs
        N = numLang = len(langs)
        heading = self.headings['administrativeData']
        hlangIdx = [i for i, elm in enumerate(heading) if elm.startswith('heading')]
        
        #%%
        # Prepare the administrativeData sheet 
        if not 'administrativeData' in wb.sheet_names:
            wb.sheets.add('administrativeData', after=after)
        sht = wb.sheets['administrativeData']
        sht.clear()
        # Write the heading
        toprow = heading
        sht.range((1,1)).value = toprow
        sht.range((1,1)).expand('right').font.bold = True
        sht.range((1,1)).expand('right').color = colors["light_gray"]
        rowIdx = 2
        
        # Get structure of the administrativeData from schema.  
        xsdAdmStruct, xsdConLocStruct = dcchf.schemaGetAdministrativeDataStructure(xsd_root,langs)
        
        # Get data from the xml and write it to rowData 2D list
        # ---------------------------------------------------------------------
        dlangIdx = {extractHeadingLang(elm):i for i, elm in enumerate(heading) if elm.startswith('heading[')}
        for rIdx, rowData in enumerate(xsdAdmStruct):
            # Sheet heading: (level, description, value, headings, xsdType, xPath)
            xsdType = rowData[-2]
            descr = rowData[1]
            xPath = rowData[-1]
            nodes = dcchf.xpath_query(root, xPath)
            # print(rowData[1], end=': ')
            if len(nodes) == 0: continue
            if "FieldType" in xsdType:
                value = nodes[0].attrib['value']
                rowData[2] = value
            if "@" in descr: 
                value = nodes[0]
                rowData[2] = value
            else:
                # add the headings from xml to rowData 
                xmlheadings = nodes[0].findall("dcx:heading", ns)
                for elm in xmlheadings:
                    lang = elm.attrib['lang']
                    if lang in langs:
                        rowData[dlangIdx[lang]] = elm.text


        # write the rowData to the sheet    
        sht.range((rowIdx,1)).value = xsdAdmStruct
        

        # set colors and indentLevel in the sheet 
        wb.app.screen_updating = False
        sectColors = [colors["gray"], colors["gray"], colors['gray']] + [colors['blue']]*N + [colors['gray']]*2
        dataColors = [colors["light_gray"], colors["light_gray"], colors['light_yellow']] + [colors['light_blue']]*N + [colors['light_gray']]*2
        for i, rowData in enumerate(xsdAdmStruct):
            sht.range((rowIdx+i,2)).api.IndentLevel = rowData[0]
            sht.range((rowIdx+i,6)).api.IndentLevel = rowData[0]
            sht.range((rowIdx+i,7)).api.IndentLevel = rowData[0]
            if rowData[0] <= 1 or rowData[-2] == 'dcx:responsiblePersonType':
                # case of a section heading
                for j,c in enumerate(sectColors):
                    sht.range((rowIdx+i,j+1)).color = c
                    sht.range((rowIdx+i,j+1)).font.bold = True
            else: 
                # case of a data row
                for j,c in enumerate(dataColors):
                    sht.range((rowIdx+i,j+1)).color = c
        
        rng = sht.range((2,3))
        rng.color = colors["light_yellow"]      

        rng = sht.range((1,1)).expand() 
        rng.api.Borders.Weight = 2

        wb.app.screen_updating = False

        #%% set validator dropdowns in sheet
        for rIdx, rowData in enumerate(xsdAdmStruct, start=2):
            xsdType = rowData[-2]
            rng = sht.range((rIdx,3))
            if xsdType == 'dcx:transactionContentFieldType':
                self.applyValidationToRange(rng, 'transactionContentType')
            elif xsdType == 'dcx:dcx:performanceLocationFieldType':
                self.applyValidationToRange(rng, 'performanceLocationType')
            elif xsdType == 'dcx:accreditationNormFieldType':
                self.applyValidationToRange(rng, 'accreditationNormType')
            elif xsdType == 'dcx:stringISO3166FieldType':
                self.applyValidationToRange(rng, 'stringISO3166Type')
            elif xsdType == 'dcx:accreditationApplicabilityFieldType':
                self.applyValidationToRange(rng, 'accreditationApplicabilityType')

        wb.app.screen_updating = True


        
        # Load and Insert contacts and Locations
        # =====================================================================
        wb.app.screen_updating = False

        # write a heading in the excel sheet 
        rIdx = rowIdx + len(xsdAdmStruct) + 5
        rng = sht.range((rIdx-1,1))
        rng.value = ['Contacts & Locations']
        rng.name = 'ContactAndLocationsStartRow'

        # Write the column heading names 
        # -----------------------------------

        #find the contact and location names in the schema
        xsdConLocNames = xsdConLocStruct[0][1:]  # ['client', 'serviceProvider', ..., 'location'] +  

        # How many dcx:locations are there in the xml-file.
        locationName = xsdConLocNames[-1] # 'location'
        # number of dcx:location nodes in the xml-file.
        locations = dcchf.xpath_query(root, f".//dcx:administrativeData//dcx:{locationName}")
        numLocationsInXml = len(locations) # allways have at least one column for location entry
        
        # construct the column heading to be entered in excel sheet. 
        tblHeading = ['level', 
                      'schemaType',
                      'description', 
                      ] \
                        + [f'heading[{lang}]' for lang in langs] \
                        + xsdConLocNames \
                        + ['location']*(numLocationsInXml) 
               
        # write to column headings to excel and apply formatting 
        sht.range((rIdx,1)).value = tblHeading
        sht.range((rIdx,1)).expand('right').font.bold = True
        sht.range((rIdx,1)).expand('right').color = colors["gray"]
        rIdx += 1

        # write the row headings, types and level info to excel sheet
        # --------------------------
        
        # get names and types: [['heading', '@id', '@imageRefs', ... ], ['dcx:stringWithLangType', 'dcx:stringWithLangType', 'xs:ID',...]]
        conLocNames, conLocTypes = dcchf.getNamesAndTypes(xsd_root,"contactAndLocationType")

        # level info is "" for elements in dcx:administrativeData, else either 'address', 'contactInfo', 'geoPosition' if in a subnode.
        conLocLevel = [""]*len(conLocNames)
        conLocSubSec = ['address', 'contactInfo', 'geoPosition']
        for tag in conLocSubSec:
            subStruct, subTypes = dcchf.getNamesAndTypes(xsd_root,tag+"Type")
            idx = conLocNames.index(tag) + 1
            conLocNames[idx:idx] = subStruct #[tag+'/'+subTag for subTag in subStruct]
            conLocTypes[idx:idx] = subTypes
            conLocLevel[idx:idx] = [tag]*len(subStruct)

        conLocStruct = list(map(list,zip(conLocLevel, conLocTypes, conLocNames)))

        # insert chosen languages for dcx:heading and dcx:body elements.  
        conLocStruct[:0:] = [['', "dcx:stringWithLangType", f"heading[{lang}]"] for lang in langs]
        idx = conLocNames.index('body')+len(langs)
        conLocStruct[idx+1:idx+1] = [['', "dcx:stringWithLangType", f"body[{lang}]"] for lang in langs]
        conLocStruct[idx][1] = "xs:string"

        # write the row heading info and data into the xl-sheet. 
        sht.range((rIdx,1)).value = conLocStruct

        # get human readable column headings and insert in the sheet. 
        # ------------------------
                    # Perhaps Delete
                    # tag = 'contactAndLocationColumnHeadingNamesType'
                    # conLocColumnHeadingNames = dcchf.schema_get_restrictions(xsd_root, tag)[tag] 

        # find the node with the human-readable column headings. 
        tblColHeadingNode = dcchf.xpath_query(root, ".//dcx:contactColumnHeadings")[0]
        # dcchf.print_node(tblColHeadingNode)

        conLocStruct = dcchf.transpose_2d_list(conLocStruct)
        conLocLevel, conLocTypes, conLocNames = conLocStruct



        # for each row get the human readable headings (different languages)
        for idx, name in enumerate(conLocNames):
            # print(name)
            if len(conLocLevel[idx]) > 0: 
                level = conLocLevel[idx]
                xsdType = conLocTypes[idx]
                name = level+'/'+name
            node = tblColHeadingNode.find(f".//dcx:column[@name='{name}']", ns)

            if node != None: 
                # find and insert the human readable row data headings to xl-sheet  
                for i, lang in enumerate(langs):
                    n = node.find(f".//*[@lang='{lang}']")
                    print("n: ",n)
                    if not n is None: 
                        # dcchf.print_node(n)
                        sht.range((rIdx+idx),(4+i)).value = n.text



        
        # find and insert contact / location data
        # --------------------------------------------------
        # sheet index of the client column
        clientIdx = tblHeading.index('client')

        # for each row get the data and insert in xl-sheet
        adminNode = dcchf.xpath_query(root, "//dcx:administrativeData")[0] # the XML dcx:administrativeData node. 
        conLocNodes = [adminNode.findall(f".//dcx:{clh}",ns) for clh in xsdConLocNames]
        for cidx, nodes in enumerate(conLocNodes): # [[Element.'client'], [Element.'serviceProvider'], ..., [Element.'location', Element.'location',...]]
            for ci, node in enumerate(nodes):
                for ridx, name in enumerate(conLocNames): # ['heading[en]', '@id', ..., body[es]]
                    # construct the correct subpath
                
                    subsec = conLocLevel[ridx]   # either ['', 'address', 'contactInfo','geoPosition']
                    s = '' if len(subsec) == 0 else f'dcx:{subsec}/'
                    if '@' in name: 
                        subPath = f'{s}{name}'
                    elif name == 'body':
                        continue
                    elif "[" in name:
                        lang = extractHeadingLang(name)
                        subPath = f"./{s}dcx:{name[:-4]}[@lang='{lang}']/text()"
                    else:
                        subPath = f'./{s}dcx:{name}/text()'
                    value = dcchf.xpath_query(node,subPath)
                    print(node, subPath, value)
                    if len(value)>0:
                        rng = sht.range((rIdx+ridx),(clientIdx+1+cidx+ci))
                        rng.value = value[0]

        
        # wb.app.screen_updating = True
        #%%
        # set cell colors of the contact and location table
        wb.app.screen_updating = False
        numConLocCols = len(xsdConLocNames) + numLocationsInXml
            
        for idx, xdsType in enumerate(conLocTypes):
            if xdsType in ['dcx:stringWithLangType']: 
                # case of column headings and bodies 
                rowcolors = [None]*(3+len(langs)) + [colors['light_blue']]*numConLocCols
                setBold = False
            elif conLocNames[idx] == 'body' or xdsType in ['dcx:addressType', 'dcx:contactInfoType', 'dcx:geoPositionType']: 
                # row body heading or case the 3 [address, contactInfo, geoPosition]
                rowcolors = [colors['gray']]*3 + [colors['blue']]*len(langs) + [colors['gray']]*numConLocCols
                setBold = True
            else: # case data element
                rowcolors = [colors['light_gray']]*3 + [colors['light_blue']]*len(langs) + [colors['light_yellow']]*numConLocCols
                setBold = False
            for j,c in enumerate(rowcolors):
                rng = sht.range((rIdx+idx,j+1))
                rng.color = c
                rng.font.bold = setBold
                rng.api.Borders.Weight = 2
                
        # set column widths
        columnWidths = [4,25,23.33]+[26]*numConLocCols
        for idx,width in enumerate(columnWidths):
            sht.range((1,idx+1)).column_width = width

        # sht.range((rowIdx+i,2)).api.IndentLevel = rowData[0]
        wb.app.screen_updating = True
        






#%%

def extractHeadingLang(s):
    r = re.findall(r'\[(.*?)\]', s)
    if len(r) == 1:
        return r[0]
    else:
        return re.findall(r'\[(.*?)\]', s)
#%%    

def removeHeadingLang(text: str):
    # Remove content between "[xx]"
    text = re.sub(r'\[\w{2}\]', '', text)
    return text

#%%

def exportToXmlFile(wb, fileName='output.xml'):
    # create an ElementMaker instance with multiple namespaces
    #%%
    myNameSpace = DCC.strip('{}')
    ns = {'dcx': 'https://dfm.dk',
            'xsi': 'http://www.w3.org/2001/XMLSchema-instance'}
    elmMaker = ElementMaker(namespace=myNameSpace, nsmap=ns)

    # create the root element of the output tree  with attributes
    sht = wb.sheets[wb.sheet_names.index('Definitions')]
    vals = sht.range((1,1)).expand('down').value
    xsdVerIdx = vals.index('SchemaVersion')+1
    xsdVer = sht.range((xsdVerIdx,2)).value
    #%%
    exportRoot = elmMaker("digitalCalibrationExchange", schemaVersion=xsdVer)
    exportRoot.set("{http://www.w3.org/2001/XMLSchema-instance}schemaLocation", myNameSpace+" dcx.xsd")

    # wb = xw.Book('DCC_pipette_blank.xlsx')
    
    adminSht = wb.sheets['administrativeData']
    global dccGuiTool 
    global HEADINGS
    adminHeading = HEADINGS['administrativeData']
    
    colIdxXpath = adminHeading.index('xPath')+1
    rowInitXpath = 2
    rngXpath = adminSht.range((rowInitXpath,colIdxXpath)).expand('down')
    rngDescription = adminSht.range((rowInitXpath,adminHeading.index('description'))).expand('down')
    descriptions = [c.value for c in rngDescription]

    def my_date_handler(year, month, day, **kwargs):
        return "%04i-%02i-%02i" % (year, month, day)

    def exportHeading(node, elmMaker, xlSheet, sheetHeading, rowIdx): 
        headingColIdxLang = [(idx, extractHeadingLang(h)) for idx,h in enumerate(sheetHeading) if h.startswith('heading[')]
        headingText = [(xlSheet.range((rowIdx,idx+1)).value, lang) for idx, lang in headingColIdxLang]
        for txt,lang in headingText:
            if not txt == None:
                elm = elmMaker("heading", txt, lang=lang)
                node.append(elm)

    def add_subtree(parent, xpath_list: list):
        valColIdx = adminHeading.index('@value')+1
        for idx, xpath in enumerate(xpath_list):
            data = adminSht.range((idx+rowInitXpath,valColIdx)).options(dates=my_date_handler).value
            isSection = adminSht.range((idx+rowInitXpath,1)).font.bold
            data = "" if data is None else str(data)
            descr = str(adminSht.range((idx+rowInitXpath, 2)).value)
            rowIdx = idx + rowInitXpath
            xpathTags = xpath.split('/')[1:]  # e.g. ['dcx:administrativeData', 'dcx:accreditation', '@statementRef']
            current_node = parent             # e.g. dcx.digitalCalibrationExchange
            lastTag = xpathTags[-1]
            for i, nodeTag in enumerate(xpathTags[:-1]):
                tag = nodeTag.replace('dcx:','')
                tag = removeHeadingLang(tag)
                # tag = tag.split("[",1)[0]
                sub_nodes = current_node.findall(nodeTag, ns)
                sub_node = sub_nodes[-1] if len(sub_nodes) > 0 else None
                if i < len(xpathTags)-1: # if not the last element
                    if sub_node == None: 
                        next_node = elmMaker(tag) 
                        current_node.append(next_node)
                        current_node = next_node
                    else:
                        current_node = sub_node
                        
            tag = lastTag
            tag = tag.replace('dcx:','')
            tag = removeHeadingLang(tag)

            if tag.startswith('@'): # case it is an attribute
                print(f'tag is: {tag}  attrib[{descr}] = {data} ', current_node)
                sub_node.attrib[descr[1:]] = data
            else:   
                if tag == "title": 
                    print(f'title is: {tag}  value: {data} ', current_node)
                    next_node = elmMaker(tag, value=data)
                    current_node.append(next_node)
                    exportHeading(next_node, elmMaker, adminSht, adminHeading, rowIdx) 
                    # next_node.attrib['value'] = data
                elif isSection: 
                    print(f'section is: {tag}  value: {data} ', current_node)
                    next_node = elmMaker(tag)
                    current_node.append(next_node)
                    exportHeading(next_node, elmMaker, adminSht, adminHeading, rowIdx) 
                # next_node.attrib['value'] = data
                elif data:
                    print(f'data is: {tag}  value: {data} ', current_node)
                    next_node = elmMaker(tag, value=data)
                    current_node.append(next_node)
                    exportHeading(next_node, elmMaker, adminSht, adminHeading, rowIdx) 
    
    s = '/dcx:digitalCalibrationExchange'
    xpath_list = [xpth.value.replace(s,'.') for xpth in rngXpath]

    add_subtree(exportRoot, xpath_list)
    adminNode = exportRoot.find('./dcx:administrativeData',ns)
    #%%
    # remove empty subnodes of administrativeData. 
    tags = ['dcxSoftware', 'accreditation', 'conformitySummary', 'coSigner', 'coSigner2', 'coSigner3','authorizingPerson', 'documentAuthorization']
    for tag in tags: 
        daNodes = adminNode.findall('.//dcx:'+tag, ns)
        for daNode in daNodes:
            if not daNode.findall('.//*[@value]'):
                daNode.getparent().remove(daNode)

    dcchf.print_node(adminNode)
    #%%

    # ToDo: add routine to remove node sections without data. 
    # use the etree remove method.


    # export Contact and Locations TABLE
    # -----------------------------------------------------
    #%%
    adminNode = exportRoot.find(".//dcx:administrativeData",ns)
    rngConLoc = adminSht.range('ContactAndLocationsStartRow')
    rngConLocHeading = rngConLoc.offset(1,0).expand('right') # Column Headings of the contact Location Table
    conLocHeading = rngConLocHeading.value # ['level', 'schemaType', 'description',  'heading[da]', 'heading[en]', 'client',
    rngXsdType = rngConLoc.offset(2,1).expand('down')
    rngDescr = rngConLoc.offset(2,2).expand('down')
    rngLevel = rngXsdType.offset(0,-1)

    global NUM_LANGS
    N = NUM_LANGS

    # export Contact And Location ColumnHeadings. 
    #-----------------------------------------
    rngConLocColHeadingLang = rngConLoc.offset(1,3).resize(1,N)
    conLocColHeading = rngConLocColHeadingLang.value
    langs = [extractHeadingLang(h) for h in conLocColHeading]

    rngConLocHeads = rngDescr.offset(0,1).resize(column_size=N)
    conLocHeads = rngConLocHeads.value
    if any(conLocHeads): 
        conLocColHeadNode = elmMaker('contactColumnHeadings')
        adminNode.append(conLocColHeadNode)

        for i, cell in enumerate(rngDescr,start=1):
            if i <= N or i > rngDescr.size - N:
                continue
            else: 
                descr = cell.value
                type = cell.offset(0,-1).value
                level = cell.offset(0,-2).value
                rngColHeadings = cell.offset(0,1).resize(1,N)
                colHeading = rngColHeadings.value
                if any(colHeading):
                    l = level+"/" if level else ""
                    conLocColNode = elmMaker('column', name=l+descr)
                    conLocColHeadNode.append(conLocColNode)
                    for h,l in zip(colHeading,langs):
                        if h: 
                            n = elmMaker('heading',h, lang=l)
                            conLocColNode.append(n)

    dcchf.print_node(adminNode)
    # adminNode.remove(adminNode.find("./dcx:contactColumnHeadings",ns))

    #%%
    # export contact and locations
    # -----------------------------
    # find the subsetion names (as written in level column). 
    subSecNames = []    #  = [None, 'address', 'contactInfo','geoPosition']
    [subSecNames.append(val) for val in rngLevel.value if val not in subSecNames]
    subSecNames.remove(None) # = ['address', 'contactInfo','geoPosition']

    # find the client column and columns to the right. 
    rngClient = rngDescr.offset(0,N+1)
    rngColumns = rngClient.expand('right')
    numConLoc = len(conLocHeading) - conLocHeading.index('client')
    for i in range(numConLoc):
        # get values from the selected column  
        rngColumn = rngClient.offset(0,i)
        colName = rngColumn.offset(-1,0).resize(1,1).value
        rngColumn.number_format = "@"
        colValues = rngColumn.options(numbers=str).value
        
        if any(colValues): # if the column has content then add it to the nodetree. 
            colNode = elmMaker(colName) 
            adminNode.append(colNode)

            subSecNodes = {None:colNode}
            for subsec, d, value in zip(rngLevel.value, rngDescr.value, colValues):
                # d = descr.value     # description value e.g. heading[en], @id, @imageRefs, name, ...
                # rowIdx = descr.row
                # subsec = adminSht.range(rowIdx,1).value # level value
                if d in subSecNames:  # if in subsectionNames = ['address', 'contactInfor', 'geoPosition']
                    n = elmMaker(d)  # create the node.
                    subSecNodes[d] = n
                    colNode.append(n)
                elif not value: 
                    continue
                elif d.startswith('@'):
                    subSecNodes[subsec].attrib[d.strip('@')] = value
                elif d.startswith('heading[') or d.startswith('body['):
                    l = extractHeadingLang(d)
                    d = removeHeadingLang(d)
                    newNode = elmMaker(d,value, lang=l) 
                    subSecNodes[subsec].append(newNode)
                else: 
                    newNode = elmMaker(d,value)
                    subSecNodes[subsec].append(newNode)
        
            for sc in subSecNames:
                n = subSecNodes[sc]
                if not n.getchildren(): 
                    colNode.remove(n)

    dcchf.print_node(adminNode)


    #%%
    # dcchf.print_node(exportRoot)
    

    # set TRUE/FALSE to true/false

    adminNode = exportRoot.find('./dcx:administrativeData', ns)
    # print(adminNode)

    #%%
    exportInfoTable(exportRoot, elmMaker, wb, nodeName = 'statementList')
    exportInfoTable(exportRoot, elmMaker, wb, nodeName = 'equipmentList')
    exportInfoTable(exportRoot, elmMaker, wb, nodeName = 'settingList')
    exportInfoTable(exportRoot, elmMaker, wb, nodeName = 'measurementConfigList')
    msIdx = wb.sheet_names.index('quantityUnitDefList')+1
    measurementResultsNode = elmMaker('measurementResultList')
    exportRoot.append(measurementResultsNode)
    for tblId in wb.sheet_names[msIdx:]:
        print(tblId)
        exportDataTable(measurementResultsNode, elmMaker, wb, tblId)
    
    quDefNode = exportInfoTable(exportRoot, elmMaker, wb, nodeName='quantityUnitDefList')
    # exportEmbeddedFiles(exportRoot, elmMaker, wb, quDefNode)

    efNode = exportInfoTable(exportRoot, elmMaker, wb, nodeName='embeddedFileList')
    exportEmbeddedFiles(exportRoot, elmMaker, wb, efNode)
    dcchf.print_node(exportRoot)

    # measurementResultsNode = elmMaker("measurementResults", name="test")

    #%% dbh
    # write the XML to file with pretty print
    with open(fileName, 'wb') as f:
        # xml_str = et.tostring(exportRoot, pretty_print=True, xml_declaration=True, encoding='utf-8').decode()
        # xml_str_crlf  = xml_str.replace('\n', '\r\n')
        # f.write(xml_str_crlf.encode())
        f.write(et.tostring(exportRoot, pretty_print=True, xml_declaration=True, encoding='utf-8'))

    dcchf.validate(fileName, 'dcx.xsd')
    #%%

def exportSheetHeading(parentNode, sht, elmMaker): 
    global NUM_LANGS
    N = NUM_LANGS+1
    for i in range(1,N): 
        head = sht.range((i,1)).value
        lang = extractHeadingLang(head)
        txt = sht.range((i,2)).value if not sht.range((i,2)).value == None else "" 
        elm = elmMaker("heading", txt, lang=lang)
        parentNode.append(elm)

def exportSheetColumnHeading(parentNode, sht, rngHeader, elmMaker):
    """
    INPUT
    --------------
        parentNode: should be the node of 'statementList', 'equipmentList', ...
        sht: xlwings sheet for the corresponding data
        rngHeader: header row of the corresponding table in the sheet. 
        elmMaker : ...
    """
    global NUM_LANGS
    N = NUM_LANGS
    headings = rngHeader.value
    rngColHeading = sht.range((N+1,2)).resize(N,len(headings)-1)
    rngColHeadingLang = sht.range((N+1,1)).resize(N,1)
    
    langs = [extractHeadingLang(h) for h in rngColHeadingLang.value]
    langByRow = {h.row: extractHeadingLang(h.value) for h in rngColHeadingLang}

    colHeads = rngColHeading.value
    if rngColHeading.shape[0] > 1: #flatten the list
        colHeads = [x for xs in colHeads for x in xs]
    if any(colHeads): 
        colHeadNode = elmMaker('columnHeadings')
        parentNode.append(colHeadNode)

        for cell in rngHeader[1:]:
            lang = extractHeadingLang(cell.value)
            tag = removeHeadingLang(cell.value)
            rngColHeadings = cell.offset(-N,0).resize(N,1)
            colHeading = rngColHeadings.value
            if any(colHeading):
                c = colHeadNode.findall(f"./*[@name='{tag}']",parentNode.nsmap)
                colNode = c[0] if len(c) > 0 else elmMaker('column', name=tag)
                colHeadNode.append(colNode)
                for h,l in zip(colHeading,langs):
                    if h: 
                        tmp = colNode.findall(f"./*[@lang='{l}']",colNode.nsmap)
                        n = tmp[0] if len(tmp) > 0 else elmMaker('heading',h, lang=l)
                        # n = elmMaker('heading',h, lang=l)
                        colNode.append(n)

    #workinghere
    #dcchf.print_node(parentNode)


def exportEmbeddedFiles(exportRoot, elmMaker, wb, embeddedFilesNode):
    ns = exportRoot.nsmap
    shtIdx = wb.sheet_names.index('embeddedFileList')
    sht = wb.sheets[shtIdx]
    
    def encode_file_to_base64(file_path):
        try:
            with open(file_path, "rb") as img_file:
                file_bytes = img_file.read()
                base64_encoded = base64.b64encode(file_bytes).decode("utf-8")
                return base64_encoded
        except FileNotFoundError:
            print(f"Error: File '{file_path}' not found.")
            return None

    tbl = sht.tables['Table_embeddedFileList']      
    rng_header = tbl.data_body_range  #xlwings function
    nrow, ncols = rng_header.shape
    print(rng_header.shape)
    dcchf.print_node(embeddedFilesNode)

    global NUM_LANGS

    for i, efn in enumerate(embeddedFilesNode.findall("./dcx:embeddedFile",ns)): 
        filepath = sht.range((NUM_LANGS*2+2+i,ncols+1)).value
        print(f"loading file: {filepath}")
        base64str = encode_file_to_base64(filepath)
        node = elmMaker('fileContent',base64str)
        efn.append(node)


def exportInfoTable(parentNode, elmMaker,wb, nodeName = 'settingList'): 
    shtIdx = wb.sheet_names.index(nodeName)
    sht = wb.sheets[shtIdx]
    statementsNode = elmMaker(nodeName)

    exportSheetHeading(statementsNode, sht, elmMaker)
    tbl = sht.tables['Table_'+nodeName]
    rng = tbl.data_body_range
    rngHeader = tbl.header_row_range
    headings = rngHeader.value
    if rng != None: 
        nrow, ncols = rng.shape
        tbl_data = rng.value
        if nrow == 1: 
            tbl_data = [tbl_data]
        for row in tbl_data:
            if row[0].startswith('y'):
                a = {h[1:]:  row[i] for i,h in enumerate(headings) if h.startswith('@')}
                a = {k:v for k,v in a.items() if v != None}
                subNodeName = nodeName.replace("List","")
                if nodeName != 'equipment':
                    node = elmMaker(subNodeName, **a)
                for i, h in enumerate(headings):
                    if i == 0 or row[i] == None: 
                        continue
                    if h.startswith('@'):
                        continue
                    elif h.startswith('heading['): 
                        lang = extractHeadingLang(h)
                        elm = elmMaker('heading', str(row[i]),lang=lang)
                    elif h.startswith('body['):
                        lang = extractHeadingLang(h)
                        elm = elmMaker('body', str(row[i]),lang=lang)
                    else: 
                        elm = elmMaker(h,str(row[i]))
                    node.append(elm)
            statementsNode.append(node)
    exportSheetColumnHeading(statementsNode, sht, rngHeader, elmMaker)
    parentNode.append(statementsNode)
    return statementsNode


def exportDataTable(parentNode, elmMaker, wb, tableId): 
    shtIdx = wb.sheet_names.index(tableId)
    sht = wb.sheets[shtIdx]
    heading = [c.value for c in sht.range((1,1)).expand('down')]
    values = [c.value for c in sht.range((1,2),(len(heading),2))]
    nodeTag = values[0]

    attrib = {}
    for i,h in enumerate(heading): 
        if h.startswith('@') and values[i] != None: 
            val = values[i]
            if h[1:].startswith('num'): 
                val = int(val)
            attrib[h[1:]] = str(val)
    # print(attrib)
    tblNode = elmMaker(nodeTag, **attrib)

    for i,h in enumerate(heading): 
        if h.startswith('heading['):
            lang = extractHeadingLang(h)
            if values[i]:
                elm = elmMaker('heading', values[i],lang=lang)
                tblNode.append(elm)
            else: 
                continue

    ncols = sht.range((len(heading)+2,1)).expand('right')
    colIdxs = range(2,len(ncols)+1)
    # print(list(colIdxs))
    for colIdx in colIdxs:
        exportDataColumn(tblNode, sht, elmMaker, wb, len(heading)+2, colIdx)

    parentNode.append(tblNode)

def exportDataColumn(parentNode, tblSheet, elmMaker, wb, rowInitIdx, colIdx): 
    typecast_dict = {'int': int, 'real': float, 'string': str, 'bool': bool, 'conformityStatus': str, 'ref': str} # Deprecated 
    numRows = int(parentNode.attrib['numRows'])
    # print('numRows = ', numRows)
    global HEADINGS
    colHeading = HEADINGS['column']
    dataInitRowIdx = rowInitIdx+len(colHeading)-1
    colAttrRange = tblSheet.range((rowInitIdx,colIdx),(dataInitRowIdx, colIdx))
    colAttrNameRange = tblSheet.range((rowInitIdx,1),(dataInitRowIdx, 1)) 
    colDataRange = tblSheet.range((dataInitRowIdx+1,colIdx),(dataInitRowIdx+numRows, colIdx)) 
    colIndexRange = tblSheet.range((dataInitRowIdx+1,1),(dataInitRowIdx+numRows, 1)) 
    
    
    colAttrNames = [c.value for c in colAttrNameRange]
    colAttrValues = [c.value for c in colAttrRange]
    colData = [c.api.Text for c in colDataRange]
    colIndex = [int(c.value) for c in colIndexRange]

    colHeadDict = dict(zip(colAttrNames,colAttrValues))

    colAttrKeys = ['scope', 'dataCategoryRef', 'quantity', 'unit', 'quantityUnitDefRef', 'settingRef'] 
    colAttrKeys = [k for k in colHeading if k.startswith("@")]
    colAttr = {k.strip('@'): colHeadDict[k] for k in colAttrKeys if colHeadDict[k]!=None}
    # colNode = elmMaker('column', **dict(zip(colAttrNames[:4], colAttrValues[:4])))
    colNode = elmMaker('column', **colAttr)

    colHeadingDict = {k: colHeadDict[k] for k in colAttrNames if k.startswith('heading')}
    for k,v in colHeadingDict.items():
        if not v == None:
            lang = extractHeadingLang(k)
            elm = elmMaker('heading', v,lang=lang)
            colNode.append(elm)

    dataCategoryNode = elmMaker(colHeadDict['dataCategory'])
    
    for i, data in enumerate(colData):
        if not data == "":
            txt = str(data)
            # print(txt)
            # elm = elmMaker(dataType, txt, idx=str(colIndex[i]))
            elm = elmMaker('row', txt, idx=str(colIndex[i]))
            dataCategoryNode.append(elm)
    # colNode.append(dataList)
    colNode.append(dataCategoryNode)
    parentNode.append(colNode)

# exportToXmlFile()    





class MainApp(tk.Tk):

    def __init__(self, guiTool: DccGuiTool):
        super().__init__()
        app = self
        self.guiTool = guiTool
        self.setup_gui(app)

        # self.configure(background='white')
        # self.bind()
        # self.bindVirtualEvents()
        uiExcelFileName = 'DCC_pipette_blank.xlsx'
        uiExcelFileName = 'DCC_UI_blank.xlsx'
        self.guiTool.loadExcelWorkbook(workBookFilePath=uiExcelFileName)
        # self.label1.config(text='DCC_pipette_blank.xlsx')
        self.label1.config(text=uiExcelFileName)
        # self.guiTool.loadDCCFile('I:\\MS\\4006-03 AI metrologi\\Software\\DCCtables\\master\\Examples\\Stip-230063-V1.xml')
        # self.label2.config(text='Stip-230063-V1.xml')
        dccFileName = 'Examples\\Stip-230063-V1.xml'
        dccFileName = 'Examples\\Template_TemperatureCal.xml'
        dccFileName = 'SKH_10112_2.xml'
        dccFileName = 'dcc-example.xml'
        # self.guiTool.loadDCCFile(dccFileName)
        #self.label2.config(text=dccFileName)
        # self.loadDCCsequence()
        # exportToXmlFile('output.xml')

    def setup_gui(self,app):
        self.wm_title("DCX EXCEL UI Tool")
        # self.canvas = tk.Canvas(self, width = 1851, height = 1041)
        self.geometry('400x200')
        self.attributes('-topmost', True)
        # img = tk.PhotoImage(file = 'BG_design.png')
        # self.background_image = img
        # self.background_label = tk.Label(app, image=img)
        # self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        
        # Create three buttons
        button1 = tk.Button(app, text='load excel-file', command=self.loadExcelBook)
        button1.pack(pady=10)
        self.label1 = tk.Label(app, text = "Select a schema file")
        self.label1.pack()

        button2 = tk.Button(app, text='load DCC.xml', command=self.loadDCC)
        button2.pack(pady=10)
        self.label2 = tk.Label(app, text = "Select a DCC file")
        self.label2.pack()

        button3 = tk.Button(app, text='Export DCC from GUI', command=self.exportDCC) #, command=self.guiTool.runDccQuery)
        button3.pack(pady=10)

    def loadDCCsequence(self):
        self.guiTool.loadDccSequence()


    def loadExcelBook(self):
        file_path = tkfd.askopenfilename(initialdir=os.getcwd())
        self.guiTool.loadExcelWorkbook(workBookFilePath=file_path)
        self.label1.config(text=file_path)


    def loadSchema(self):
        file_path = tkfd.askopenfilename(initialdir=os.getcwd())
        self.guiTool.loadSchemaFile(xsdFileName=file_path)
        self.guiTool.loadSchemaRestrictions()
        self.label1.config(text=file_path)

    def loadDCC(self):
        file_path = tkfd.askopenfilename(initialdir=os.getcwd())
        errors = self.guiTool.loadDCCFile(file_path)
        if not errors: 
            self.label2.config(text=file_path)
        else: 
            self.label2.config(text="File does not validate!")
        
    def exportDCC(self):
        # file_path = tkfd.asksaveasfilename(initialdir=os.getcwd())
        file_path = "output.xml"
        self.label2.config(text="EXPORTING!")
        exportToXmlFile(self.guiTool.wb, file_path)
        self.label2.config(text=f"exported to: {file_path}")
        
class MyLanguageDialog:
    def __init__(self, parent, languages):
        top = self.top = tk.Toplevel(parent)
        top.geometry('300x200')
        top.attributes('-topmost', True)
        self.myLabel = tk.Label(top, text='Select Language')
        self.myLabel.grid(row=0,column=0, columnspan=2)

        numLangInDCC = languages.index('---')

        N = 6
        self.langs = [tk.StringVar() for k in range(N)]
        lbls = ['Mandatory lang']+[f'Used lang {i}' for i in range(1,N)]
        self.myLabels = [tk.Label(top, text=s) for s in lbls]
        self.tkCboxs = [ttk.Combobox(top, textvariable=self.langs[k]) for k in range(N)]

        for k in range(N): 
            self.tkCboxs[k].grid(row=k+1, column=1)
            self.tkCboxs[k]['values'] = languages
            i = min(k,numLangInDCC)
            self.langs[k].set(languages[i])
            self.myLabels[k].grid(row=k+1, column=0)
       
        self.mySubmitButton = tk.Button(top, text='Submit', command=self.send)
        self.mySubmitButton.grid(row=N+2, column=0, columnspan=2)

    def send(self):
        global dccGuiTool
        langs = [l.get() for l in self.langs if l.get()!="---"]
        global NUM_LANGS
        dccGuiTool.langs = langs
        print("Selected Languages: ", dccGuiTool.langs)
        NUM_LANGS = len(langs)
        print(NUM_LANGS)
        self.top.destroy()


if __name__=="__main__":
    #################
    #first argument is dcc xml file
    #second argument is excel template to use
    import sys
    args=sys.argv[1:]
    print(len(args))
    

    dccGuiTool = DccGuiTool()
    app = MainApp(dccGuiTool)
    app.mainloop()
    # if len(args)==0:
    #     mapFileName ='Examples'+os.sep+'Mapping_Novo_temperatur_Certifikat.xlsx'
    #     dccFileName = 'Examples'+os.sep+'Stip-230063-V1.xml'
    #     lookupFromMappingFile(mapFileName, dccFileName)
    # elif len(args)==2:
    #     mapFileName = args[0]
    #     dccFileName = args[1]
    #     lookupFromMappingFile(mapFileName, dccFileName)
    # else: 
    #     helpstatement = """call dccquery.py using the following arguments: \n 
    #     >> python dccquery.py [mapping file e.g. mapping.xlsx] [DCC file e.g. dcc.xml] """
    #     print(helpstatement)

         

# %%
