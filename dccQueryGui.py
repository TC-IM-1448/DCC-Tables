#%%
import os
import openpyxl as pyxl
import xlwings as xw
from  lxml import etree as et
import tkinter as tk
import tkinter.filedialog as tkfd
import DCChelpfunctions as dcchf
from DCChelpfunctions import search

#%%
LANG='da'
DCC='{https://dfm.dk}'
xlValidateList = xw.constants.DVType.xlValidateList

#%%
def lookupFromMappingFile(mapFileName:str, dccFileName:str):
    """LookupFromMappingFile """
    pass


class DccQuerryTool(): 
    mapperHeading =['Client DB ref', 'client description', 'queryType', 
                    'xpath', 
                    'measurementConfigRef', 'serviceCategory', 'tableId', 
                    'scope', 'dataCategory', 'dataCategoryRef', 'quantity', 'unit', 'quantityUnitDefRef', 'idx', 
                    'query result']
    xsdDefInitCol = 3  
    def loadExcelWorkbook(self, workBookFilePath: str):
        self.wb = xw.Book(workBookFilePath)
        wb = xw.Book(workBookFilePath)
        self.sheetDef = wb.sheets['Definitions']
        self.sheetMap = wb.sheets['Mapping']
        self.wb = wb
        self.sheetMap.range((1,1)).value = self.mapperHeading
        self.loadSchemaFile()
        self.loadSchemaRestrictions()
        wb.activate(steal_focus=True)

    def loadSchemaFile(self, xsdFileName="dcx.xsd"):
        self.xsdTree, self.xsdRoot  = dcchf.load_xml(xsdFileName)
        
    def loadDCCFile(self, xmlFileName="dcc-example.xml"):
        self.dccTree, self.dccRoot = dcchf.load_xml(xmlFileName)
        
    def loadSchemaRestrictions(self): 
        xsd_root = self.xsdRoot
        sht_def = self.sheetDef
        drestr = dcchf.schema_get_restrictions(xsd_root)
        j = self.xsdDefInitCol
        for i, (k,vs) in enumerate(drestr.items()):
            sht_def.range((1, i+j)).value = [k]
            sht_def.range((2, i+j)).value = [[v] for v in vs]
            sht_def.range((2,i+j)).expand('down').name = k  
        self.dccDefInitCol = j+i+1
        
        j = self.mapperHeading.index('scope')+1

        # Apply validators in the mapping sheet
        sht = self.sheetMap
        for r in ['scopeType','dataCategoryType', 'dataCategoryType', 'quantityType']: 
            rng = sht.range((2,j),(1024,j))
            rng.api.Validation.Delete()
            fml = "="+r
            rng.api.Validation.Add(Type=xlValidateList, Formula1=fml)
            j += 1


    def loadDccAttributes(self):
        root = self.dccRoot
        sht_def = self.sheetDef
        sht_map = self.sheetMap

        i = self.dccDefInitCol
        statementIds = [elm.attrib['id'] for elm in dcchf.get_statements(root)]
        sht_def.range((1,i)).value = 'statementId'
        sht_def.range((2,i)).value = [[s] for s in statementIds]
        sht_def.range((2,i)).expand('down').name = 'statementId' 
        
        i+=1 
        msucIds = [elm.attrib['id'] for elm in dcchf.get_measurementConfigs(root)]
        sht_def.range((1,i)).value = 'measurementConfigId'
        sht_def.range((2,i)).value = [[ms] for ms in msucIds]
        sht_def.range((2,i)).expand('down').name = 'measurementConfigId'
        cidx = self.mapperHeading.index('measurementConfigRef')+1
        rng = sht_map.range((2,cidx),(1024,cidx))
        rng.api.Validation.Delete()
        rng.api.Validation.Add(Type=xlValidateList, Formula1='=measurementConfigId')

        cidx = self.mapperHeading.index('serviceCategory')+1
        rng = sht_map.range((2,cidx),(1024,cidx))
        rng.api.Validation.Delete()
        rng.api.Validation.Add(Type=xlValidateList, Formula1='=serviceCategoryType') 

        i+=1
        tableIds = [elm.attrib['tableId'] for elm in dcchf.getTables(root)]
        sht_def.range((1,i)).value = 'tableId'
        sht_def.range((2,i)).value = [[tbl] for tbl in tableIds]
        sht_def.range((2,i)).expand('down').name = 'tableId'
        cidx = self.mapperHeading.index('tableId')+1
        rng = sht_map.range((2,cidx),(1024,cidx))
        rng.api.Validation.Delete()
        rng.api.Validation.Add(Type=xlValidateList, Formula1='=tableId') 

        i+=1
        units = ["'"+u for u in dcchf.xpath_query(root,'*//dcx:column/@unit')]
        units = list(set(units))
        sht_def.range((1,i)).value = 'units'
        sht_def.range((2,i)).value = [[tbl] for tbl in units]
        sht_def.range((2,i)).expand('down').name = 'units'
        cidx = self.mapperHeading.index('unit')+1
        rng = sht_map.range((2,cidx),(1024,cidx))
        rng.api.Validation.Delete()
        rng.api.Validation.Add(Type=xlValidateList, Formula1='=units') 

        i+=1
        quDefs = [elm.attrib['id'] for elm in root.findall("*//dcx:quantityUnitDef",root.nsmap)]
        sht_def.range((1,i)).value = 'quantityUnitsDefs'
        sht_def.range((2,i)).value = [[tbl] for tbl in quDefs]
        sht_def.range((2,i)).expand('down').name = 'quantityUnitDefRefs'
        cidx = self.mapperHeading.index('quantityUnitDefRef')+1
        rng = sht_map.range((2,cidx),(1024,cidx))
        rng.api.Validation.Delete()
        rng.api.Validation.Add(Type=xlValidateList, Formula1='=quantityUnitDefRefs') 

        sht_map.range((1,2)).value = [[v] for v in ['queryType', 'data', 'xpath']]
        sht_def.range((2,2)).expand('down').name = 'queryType'
        cidx = self.mapperHeading.index('queryType')+1
        rng = sht_map.range((2,cidx),(1024,cidx))
        rng.api.Validation.Delete()
        rng.api.Validation.Add(Type=xlValidateList, Formula1='=queryType') 


        
    def runDccQuery(self):
        sht = self.sheetMap
        root = self.dccRoot
        vals = sht.range("A1").expand("down").value
        [print(v) for v in vals]
        n_rows = vals.index("--END--")+1
        # queryParam = sht.range((2,1),(n_rows-1,len(self.mapperHeading))).value
        print(n_rows)

        for i in range(2, n_rows+1):
            queryType = sht.range((i,3)).value
            cidxQueryResult = self.mapperHeading.index('query result')+1
            data_val = None
            if queryType == 'xpath':
                xpath_str = sht.range((i,4)).value
                # val = dcchf.xpath_query(root, xpath_str)
                val = root.xpath(xpath_str, namespaces=root.nsmap)
                print(vals[i-2], queryType, val)
                if len(val)>0:
                    data_val = [val]
                else:
                    data_val = "ERROR not Found"
            elif queryType == 'data':
                cidx = self.mapperHeading.index("measurementConfigRef")+1
                dtbl = dict(zip(["measurementConfigRef", "serviceCategory", "tableId"], sht.range((i,cidx),(i,cidx+3)).value))
                dtbl = {k:v for k,v in dtbl.items() if v != None}
                dcol = dict(zip(["scope", "dataCategoryRef", "measurand"], [sht.range((i,cidx+3)).value]+sht.range((i,cidx+3+2), (i,cidx+3+4)).value))
                dcol = {k:v for k,v in dcol.items() if v != None}
                dataCategory = sht.range((i,cidx+3+1)).value

                cidxUnit = self.mapperHeading.index('unit')+1
                unit = sht.range((i,cidxUnit)).value
                customerTag = sht.range((i,cidxUnit+1)).value
                if customerTag is None:
                    data = search(root, dtbl, dcol, dataCategory, unit)
                else: 
                    data = search(root, dtbl, dcol, dataCategory, unit, rowTags=[customerTag])
                print(dtbl, dcol, dataCategory, unit, customerTag, ":", data)
                if len(data) == 0 : 
                    data_val = "ERROR not Found"
                else:
                    data_val = list(data.values())
                # elif len(data) > 1: 
                    # data_val = "ERROR too many values"
            sht.range((i,cidxQueryResult)).expand('right').clear_contents()
            sht.range((i,cidxQueryResult)).value = data_val
                # sht[f"M{i}"].value = data_val





class MainApp(tk.Tk):
    def __init__(self, queryTool: DccQuerryTool):
        super().__init__()
        app = self
        self.queryTool = queryTool
        self.setup_gui(app)

        # self.configure(background='white')
        # self.bind()
        # self.bindVirtualEvents()
        
        
        
    def setup_gui(self,app):
        self.wm_title("DCC EXCEL Mapping Tool")
        # self.canvas = tk.Canvas(self, width = 1851, height = 1041)
        self.geometry('500x200')
        self.attributes('-topmost', True)
        # img = tk.PhotoImage(file = 'BG_design.png')
        # self.background_image = img
        # self.background_label = tk.Label(app, image=img)
        # self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        
        # Create three buttons
        button1 = tk.Button(app, text='load query excel-file', command=self.loadExcelBook)
        button1.pack(pady=10)
        self.label1 = tk.Label(app, text = "Select a schema file")
        self.label1.pack()

        button2 = tk.Button(app, text='load DCC.xml', command=self.loadDCC)
        button2.pack(pady=10)
        self.label2 = tk.Label(app, text = "Select a DCC file")
        self.label2.pack()

        button3 = tk.Button(app, text='run query', command=self.queryTool.runDccQuery)
        button3.pack(pady=10)

        self.startupTest()

    def startupTest(self):
        excelFileName = 'SKH_10112_2_Mapping.xlsx'
        self.queryTool.loadExcelWorkbook(excelFileName)
        self.label1.config(text=excelFileName)
        self.queryTool.loadSchemaFile(xsdFileName='dcx.xsd')
        self.queryTool.loadSchemaRestrictions()
        self.queryTool.loadDCCFile('dcc-example.xml')
        self.queryTool.loadDccAttributes()
        self.label2.config(text='dcc-example.xml')

    def loadExcelBook(self):
        file_path = tkfd.askopenfilename(initialdir=os.getcwd())
        self.queryTool.loadExcelWorkbook(workBookFilePath=file_path)
        self.label1.config(text=file_path)


    def loadSchema(self):
        file_path = tkfd.askopenfilename(initialdir=os.getcwd())
        self.queryTool.loadShemaFile(xsdFileName=file_path)
        self.queryTool.loadSchemaRestrictions()
        self.label1.config(text=file_path)

    def loadDCC(self):
        file_path = tkfd.askopenfilename(initialdir=os.getcwd())
        self.queryTool.loadDCCFile(file_path)
        self.queryTool.loadDccAttributes()
        self.label2.config(text=file_path)
        
# if __name__ == "__main__":

if __name__=="__main__":
    #################
    #first argument is dcc xml file
    #second argument is excel template to use
    import sys
    args=sys.argv[1:]
    print(len(args))
    

    dccQuerryTool = DccQuerryTool()
    app = MainApp(dccQuerryTool)
    app.mainloop()
    # if len(args)==0:
    #     mapFileName ='Examples'+os.sep+'Mapping_Novo_temperatur_Certifikat.xlsx'
    #     dccFileName = 'Examples'+os.sep+'Stip-230063-V1.xml'
    #     lookupFromMappingFile(mapFileName, dccFileName)
    # elif len(args)==2:
    #     mapFileName = args[0]
    #     dccFileName = args[1]
    #     lookupFromMappingFile(mapFileName, dccFileName)
    # else: 
    #     helpstatement = """call dccquery.py using the following arguments: \n 
    #     >> python dccquery.py [mapping file e.g. mapping.xlsx] [DCC file e.g. dcc.xml] """
    #     print(helpstatement)

         
